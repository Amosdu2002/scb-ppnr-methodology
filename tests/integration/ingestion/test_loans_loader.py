"""Workbook binding for the Corporate loan inputs.

Every fixture here is a synthetic workbook built in the test — no company data.
The point is the binding: header geometry, the four different unit scales, the
workbook's own missing-value tokens, and the two sheets whose layout the model
depends on."""

from __future__ import annotations

import datetime as _dt
from pathlib import Path

import openpyxl
import pytest

from scb_ppnr.core.schemas import ValidationFailure
from scb_ppnr.ingestion.loans_loader import (
    LoansSheetSpec,
    load_3m_treasury,
    load_category_balances,
    load_facilities,
    load_merged_bucket_balance,
)
from scb_ppnr.interest_income.loans_schemas import VT_DO_NOT_USE, VT_FIXED, VT_FLOATING

_VARIABILITY_HEADER = "Interest Rate Variability"

H1_HEADERS = [
    "Customer ID", "Line Reported on FR Y9C", _VARIABILITY_HEADER,
    "Lower of Cost or Market Flag", "Interest Rate", "Committed Exposure Global",
    "Utilized Exposure Global", "Interest Rate Floor", "Origination Date", "Maturity Date",
]


def _workbook(tmp_path: Path, h1_rows, fry9c_rows=None, mev_rows=None,
              variability_header="Interest Rate Variability", h1_headers=None) -> Path:
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "CORP H.1"
    headers = [variability_header if h == _VARIABILITY_HEADER else h
               for h in (h1_headers if h1_headers is not None else H1_HEADERS)]
    for _ in range(3):                       # headers sit on row 4
        sheet.append([None] * len(headers))
    sheet.append(headers)
    for row in h1_rows:
        sheet.append(row)

    fry9c = book.create_sheet("FR-Y9C 4Q 2024")
    for _ in range(7):                       # header row 8
        fry9c.append([None, None, None])
    fry9c.append(["ID_RSSD", "Description", "Value"])
    for row in fry9c_rows if fry9c_rows is not None else [
        ["BHCK1545", "Loans for purchasing or carrying securities", 1_500_000.0],
        ["BHDM1420", "Loans secured by farmland", 500_000.0],
    ]:
        fry9c.append(row)

    mev = book.create_sheet("MEV")
    mev.append(["Scenario Name", "Date", "3-month Treasury rate"])
    for row in mev_rows if mev_rows is not None else [
        ["Actual", "1976 Q1", 4.9],
        ["Actual", "2020 Q1", 1.5],
        ["Actual", "2022 Q2", 0.8],
        ["Actual", "2024 Q4", 4.4],
        ["Severely Adverse", "2025 Q1", 3.0],
        ["Severely Adverse", "2025 Q2", 1.8],
        ["Severely Adverse", "2025 Q3", 0.1],
    ]:
        mev.append(row)

    path = tmp_path / "corp.xlsx"
    book.save(path)
    return path


def _row(fid, code, var, locom, rate=0.06, committed=1_000_000.0, utilized=None,
         floor=None, originated="15-Oct-2024", matures="15-Oct-2026"):
    return [fid, code, var, locom, rate, committed,
            committed if utilized is None else utilized, floor, originated, matures]


# --- CORP H.1 -------------------------------------------------------------

def test_headers_are_read_from_row_four_and_scales_applied(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3, rate=0.075, committed=2_500_000.0)])
    facilities, census = load_facilities(LoansSheetSpec(workbook=path))

    assert census.rows_read == 1
    facility = facilities[0]
    assert facility.facility_id == "F1"
    assert facility.interest_rate == pytest.approx(0.075)         # already decimal
    assert facility.committed_exposure == pytest.approx(2.5)      # dollars -> millions
    assert facility.origination_date == _dt.date(2024, 10, 15)    # 15-Oct-2024


def test_the_two_collapses_survive_the_round_trip(tmp_path):
    """Codes 4 and 5 into one C&I segment; LOCOM 1 and 2 into one FVO/HFS class."""
    path = _workbook(tmp_path, [
        _row("F1", 4, 2, 3), _row("F2", 5, 2, 3),      # both C&I / HFI / Floating
        _row("F3", 4, 2, 1), _row("F4", 4, 2, 2),      # both C&I / FVO_HFS / Floating
    ])
    facilities, census = load_facilities(LoansSheetSpec(workbook=path))

    assert facilities[0].segment == facilities[1].segment
    assert facilities[2].segment == facilities[3].segment
    assert facilities[0].segment != facilities[2].segment
    # the census keeps the workbook's own key spelling, which does NOT collapse
    assert set(census.reference_keys) == {"4_2_3", "5_2_3", "4_2_1", "4_2_2"}


def test_the_null_token_is_do_not_use_and_shows_in_the_census(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, "[NULL]", 3, rate=None)])
    facilities, census = load_facilities(LoansSheetSpec(workbook=path))

    assert facilities[0].segment.variable_type == VT_DO_NOT_USE
    assert census.reference_keys["4_[NULL]_3"] == 1


@pytest.mark.parametrize("token", ["NA", "[NULL]", "NONE", "#VALUE!", ""])
def test_missing_tokens_become_none_and_are_counted(tmp_path, token):
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3, rate=token, floor=token)])
    facilities, census = load_facilities(LoansSheetSpec(workbook=path))

    assert facilities[0].interest_rate is None
    assert facilities[0].interest_rate_floor is None
    assert census.missing_interest_rate == 1
    assert census.missing_interest_rate_exposure == pytest.approx(1.0)


def test_a_populated_zero_floor_is_kept_not_treated_as_absent(tmp_path):
    """The PID-SEC-18 lesson: a real 0 and 'no floor on file' are different."""
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3, floor=0.0)])
    facilities, census = load_facilities(LoansSheetSpec(workbook=path))

    assert facilities[0].interest_rate_floor == pytest.approx(0.0)
    assert census.missing_floor == 0


def test_an_unidentified_row_is_labeled_and_kept_with_its_balances(tmp_path):
    """PID-LOAN-12: rows with [NULL] in every ID column still carry real
    balances — dropping them would silently understate segment shares, pool
    rates, and wt denominators. They get a synthesized row label instead."""
    path = _workbook(tmp_path, [
        _row("F1", 4, 1, 3, committed=1_000_000.0),
        _row("[NULL]", 4, 2, 3, committed=2_000_000.0),
    ])
    facilities, census = load_facilities(LoansSheetSpec(workbook=path))

    assert len(facilities) == 2                              # kept, not dropped
    labeled = facilities[1]
    assert labeled.facility_id == "UNIDENTIFIED-ROW-6"       # header row 4 -> data rows 5, 6
    assert labeled.committed_exposure == pytest.approx(2.0)  # balances intact
    assert census.id_sources == {"customer_id": 1, "synthesized": 1}
    assert census.unidentified_rows == [6]
    assert census.unidentified_exposure == pytest.approx(2.0)
    assert "unidentified rows (labeled) : 1" in census.render()
    # the fallback columns are absent from this fixture, so the census says so
    assert any("col_internal_id" in note for note in census.warnings)


def test_the_id_chain_prefers_real_identifiers_over_synthesis(tmp_path):
    """Customer ID -> Internal ID -> Original Internal ID -> synthesized."""
    headers = H1_HEADERS + ["Internal ID", "Original Internal ID"]
    rows = [
        _row("CUST-1", 4, 1, 3) + ["INT-1", "ORIG-1"],       # customer wins
        _row("[NULL]", 4, 2, 3) + ["INT-2", "ORIG-2"],       # internal next
        _row("[NULL]", 4, 2, 3) + ["[NULL]", "ORIG-3"],      # then original
        _row("[NULL]", 4, 2, 3) + ["[NULL]", "[NULL]"],      # then the label
    ]
    path = _workbook(tmp_path, rows, h1_headers=headers)
    facilities, census = load_facilities(LoansSheetSpec(workbook=path))

    assert [f.facility_id for f in facilities] == [
        "CUST-1", "INT-2", "ORIG-3", "UNIDENTIFIED-ROW-8",
    ]
    assert census.id_sources == {
        "customer_id": 1, "internal_id": 1, "original_internal_id": 1, "synthesized": 1,
    }
    # fallback columns exist here, so no configuration hint is emitted
    assert not any("col_internal_id" in note for note in census.warnings)


def test_a_missing_column_names_what_is_available(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)])
    spec = LoansSheetSpec(workbook=path, col_committed="Committed Exposure")
    with pytest.raises(ValidationFailure, match="available:"):
        load_facilities(spec)


def test_the_sheets_spelling_is_primary_and_needs_no_substitution(tmp_path):
    """User-confirmed 2026-08-07: the workbook spells it correctly."""
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)],
                     variability_header="Interest Rate Variability")
    facilities, census = load_facilities(LoansSheetSpec(workbook=path))

    assert facilities[0].segment.variable_type == VT_FIXED
    assert census.column_substitutions == []


def test_a_misspelled_variant_still_loads_but_is_reported(tmp_path):
    """An extract that does carry the typo is not worth a failed run — but the
    substitution is named rather than applied silently."""
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)],
                     variability_header="Interest Rate Variablility")
    facilities, census = load_facilities(LoansSheetSpec(workbook=path))

    assert facilities[0].segment.variable_type == VT_FIXED
    assert any("Variablility" in note for note in census.column_substitutions)


def test_blank_rows_are_skipped_and_the_census_renders(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3), [None] * 10, _row("F2", 8, 2, 3)])
    facilities, census = load_facilities(LoansSheetSpec(workbook=path))

    assert len(facilities) == 2
    assert "rows read                   : 2" in census.render()


# --- FR Y-9C merged bucket (PID-LOAN-10) ----------------------------------

def test_merged_bucket_sums_both_mdrms_and_converts_from_thousands(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)])
    total, parts = load_merged_bucket_balance(LoansSheetSpec(workbook=path))

    assert parts["BHCK1545"] == pytest.approx(1_500.0)   # thousands -> millions
    assert parts["BHDM1420"] == pytest.approx(500.0)
    assert total == pytest.approx(2_000.0)


def test_a_missing_mdrm_is_refused_rather_than_read_as_zero(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)],
                     fry9c_rows=[["BHCK1545", "securities", 1_500_000.0]])
    with pytest.raises(ValidationFailure, match=r"BHDM1420"):
        load_merged_bucket_balance(LoansSheetSpec(workbook=path))


# --- MEV ------------------------------------------------------------------

def test_history_and_projection_split_by_scenario_with_percent_normalized(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)])
    history, projection, launch = load_3m_treasury(
        LoansSheetSpec(workbook=path), "Severely Adverse", (1, 2, 3), "2024Q4"
    )

    assert history["1976Q1"] == pytest.approx(0.049)     # percent -> decimal
    assert history["2022Q2"] == pytest.approx(0.008)
    assert launch == pytest.approx(0.044)
    assert projection == {1: pytest.approx(0.030), 2: pytest.approx(0.018), 3: pytest.approx(0.001)}


def test_quarter_labels_lose_the_space(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)])
    history, _, _ = load_3m_treasury(
        LoansSheetSpec(workbook=path), "Severely Adverse", (1, 2, 3), "2024Q4"
    )
    assert all(" " not in label for label in history)


def test_a_short_scenario_path_names_the_missing_quarters(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)])
    with pytest.raises(ValidationFailure, match=r"missing PQ4 \(2025Q4\)"):
        load_3m_treasury(
            LoansSheetSpec(workbook=path), "Severely Adverse", tuple(range(1, 10)), "2024Q4"
        )


def test_a_misspelled_scenario_name_says_what_it_found(tmp_path):
    """The real sheet's projection block is named 'Supervisory Severely Adverse';
    a near-miss name should point at the spelling, not just fail."""
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)])
    with pytest.raises(ValidationFailure, match="scenario rows found: <none>.*spelling"):
        load_3m_treasury(
            LoansSheetSpec(workbook=path), "Severly Adverse", (1, 2, 3), "2024Q4"
        )


def test_projection_rows_map_by_date_not_sheet_order(tmp_path):
    """A scenario block sorted oddly, or carrying a tail beyond the horizon,
    must land each value on the quarter its Date names."""
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)], mev_rows=[
        ["Actual", "2024 Q4", 4.4],
        ["Supervisory Severely Adverse", "2025 Q3", 0.1],     # out of order
        ["Supervisory Severely Adverse", "2025 Q1", 3.0],
        ["Supervisory Severely Adverse", "2025 Q2", 1.8],
        ["Supervisory Severely Adverse", "2026 Q1", 9.9],     # beyond the 3Q horizon
    ])
    _, projection, _ = load_3m_treasury(
        LoansSheetSpec(workbook=path), "Supervisory Severely Adverse", (1, 2, 3), "2024Q4"
    )
    assert projection == {1: pytest.approx(0.030), 2: pytest.approx(0.018), 3: pytest.approx(0.001)}


def test_a_launch_point_absent_from_history_is_refused(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)])
    with pytest.raises(ValidationFailure, match="launch point"):
        load_3m_treasury(LoansSheetSpec(workbook=path), "Severely Adverse", (1, 2, 3), "2023Q4")


# --- workbook -> income, end to end ---------------------------------------

def test_workbook_through_to_projected_income(tmp_path):
    """The whole chain on a synthetic workbook: read, decode, launch point,
    projection — including the merged 9/10/11 bucket priced off the depository
    slice."""
    from scb_ppnr.ingestion.loans_mapping import (
        DEPOSITORY_INSTITUTION_H1_CODES,
        FED_CATEGORY_NAMES,
        scalars_by_category_name,
    )
    from scb_ppnr.interest_income.loans_launchpoint import (
        build_launch_point,
        merged_bucket_launch_point,
    )
    from scb_ppnr.interest_income.loans_projection import project_corporate

    quarters = (1, 2, 3)
    path = _workbook(tmp_path, [
        _row("C1", 4, 2, 3, rate=0.061),                       # C&I / HFI / Floating
        _row("C2", 5, 1, 3, rate=0.070, originated="15-Jan-2020",
             matures="15-Feb-2025"),                           # C&I / HFI / Fixed
        _row("D1", 1, 2, 3, rate=0.055),                       # depository, Floating
        _row("D2", 7, 2, 3, rate=0.090),                       # NONdepository, Floating
        _row("X1", 8, "[NULL]", 3, rate=None),                 # earns nothing
    ])
    spec = LoansSheetSpec(workbook=path)

    facilities, load_census = load_facilities(spec)
    merged_balance, _ = load_merged_bucket_balance(spec)
    history, projection_3m, launch_3m = load_3m_treasury(
        spec, "Severely Adverse", quarters, "2024Q4"
    )

    def quarter_of_maturity(when):
        index = (when.year - 2025) * 4 + (when.month - 1) // 3 + 1
        return index if 1 <= index <= len(quarters) else None

    balances = {name: 1000.0 for name in FED_CATEGORY_NAMES.values()}
    launch, _ = build_launch_point(
        facilities, balances, launch_3m, history, quarters, quarter_of_maturity
    )
    merged = merged_bucket_launch_point(
        facilities, merged_balance, launch_3m, DEPOSITORY_INSTITUTION_H1_CODES,
        FED_CATEGORY_NAMES[9],
    )

    scalars, _ = scalars_by_category_name()
    _, totals, _ = project_corporate(
        {**dict(launch), merged.segment: merged}, projection_3m, quarters, scalars
    )

    assert load_census.rows_read == 5
    assert totals["Commercial and industrial"][1] > 0.0
    # the merged bucket borrows ONLY the depository rate (5.5%), not the
    # nondepository row's 9.0% — both sit in the same Fed Category
    assert merged.spread.pool_rate == pytest.approx(0.055)
    assert merged.spread.spread == pytest.approx(0.055 - 0.044)
    assert merged.balance == pytest.approx(2_000.0)


def test_merged_bucket_refuses_to_default_a_borrowed_spread(tmp_path):
    from scb_ppnr.ingestion.loans_mapping import DEPOSITORY_INSTITUTION_H1_CODES
    from scb_ppnr.interest_income.loans_launchpoint import merged_bucket_launch_point

    path = _workbook(tmp_path, [_row("D1", 7, 2, 3, rate=0.09)])   # nondepository only
    facilities, _ = load_facilities(LoansSheetSpec(workbook=path))

    with pytest.raises(ValidationFailure, match="no such rows carry an interest rate"):
        merged_bucket_launch_point(
            facilities, 100.0, 0.044, DEPOSITORY_INSTITUTION_H1_CODES, "merged"
        )


# --- M.1 Balance: the sheet carries its own role wiring -------------------

def _m1_workbook(tmp_path: Path, m1_rows) -> Path:
    """M.1 layout: col A domestic role, col B international role, col C the
    FR Y-9C line, then MDRM/value pairs for domestic HFI, domestic HFS/FVO,
    international HFI, international HFS/FVO."""
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)])
    book = openpyxl.load_workbook(path)
    sheet = book.create_sheet("M.1 Balance")
    for _ in range(10):                      # data starts on row 11
        sheet.append([None] * 11)
    for row in m1_rows:
        sheet.append(row)
    book.save(path)
    return path


def _m1_row(dom_role, int_role, dom_hfi=None, dom_hfs=None, int_hfi=None, int_hfs=None):
    return [dom_role, int_role, "some FR Y-9C line",
            "MDRM", dom_hfi, "MDRM", dom_hfs, "MDRM", int_hfi, "MDRM", int_hfs]


def test_category_balances_come_from_the_sheets_own_role_columns(tmp_path):
    path = _m1_workbook(tmp_path, [
        _m1_row("Wholesale - Corp - C&I and others", "Wholesale - Corp - C&I and others",
                dom_hfi=135_038.0, dom_hfs=3_441.0, int_hfi=31_956.0, int_hfs=3_848.0),
    ])
    balances, _, census = load_category_balances(LoansSheetSpec(workbook=path))

    assert balances["Commercial and industrial"] == pytest.approx(174_283.0)
    assert "LOANS LOADER CENSUS" in census.render()


def test_one_fr_y9c_line_can_feed_two_categories(tmp_path):
    """'c. Secured by farmland' sends its domestic balance to Domestic farmland
    and its international balance to International farmland — which is exactly
    what the two role columns are for."""
    path = _m1_workbook(tmp_path, [
        _m1_row("Wholesale - Corp - farmland", "Wholesale - Corp - int farmland",
                dom_hfi=130.0, dom_hfs=0.0, int_hfi=721.0),
    ])
    balances, _, _ = load_category_balances(LoansSheetSpec(workbook=path))

    assert balances["Domestic farmland"] == pytest.approx(130.0)
    assert balances["International farmland"] == pytest.approx(721.0)


def test_retail_and_cre_rows_are_skipped(tmp_path):
    path = _m1_workbook(tmp_path, [
        _m1_row("Retail - mortgage - first lien", "Retail - noncore", dom_hfi=291_627.0),
        _m1_row("Wholesale - CRE - construction", "Wholesale - CRE - international",
                dom_hfi=13_656.0),
        _m1_row("Wholesale - Corp - agricultural", "Wholesale - Corp - agricultural",
                dom_hfi=114.0),
    ])
    balances, _, census = load_category_balances(LoansSheetSpec(workbook=path))

    assert set(balances) == {"Agriculture Loans"}
    assert balances["Agriculture Loans"] == pytest.approx(114.0)
    assert any("non-Corporate role cells" in note for note in census.warnings)


def test_an_unrecognized_corporate_role_is_refused_not_dropped(tmp_path):
    """The labels are truncated by column width in the workbook's display, so a
    role that does not match means the transcription is wrong — and dropping it
    would silently zero that category's balance."""
    path = _m1_workbook(tmp_path, [
        _m1_row("Wholesale - Corp - something new", None, dom_hfi=100.0),
    ])
    with pytest.raises(ValidationFailure, match="not one of"):
        load_category_balances(LoansSheetSpec(workbook=path))


def test_categories_absent_from_m1_are_warned_about(tmp_path):
    path = _m1_workbook(tmp_path, [
        _m1_row("Wholesale - Corp - agricultural", None, dom_hfi=114.0),
    ])
    _, _, census = load_category_balances(LoansSheetSpec(workbook=path))
    assert any("project zero income" in note for note in census.warnings)


# --- reference results (compare mode) --------------------------------------

def _results_workbook(tmp_path: Path) -> Path:
    """A results sheet in the observed layout: block markers, Fixed/Variable/
    Total rows, PQ0..PQ9 columns, raw dollars, with the Total exceeding
    Fixed + Variable by a hidden third stream."""
    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)])
    book = openpyxl.load_workbook(path)
    sheet = book.create_sheet("Results")
    sheet.append([None, None, None] + [f"PQ{q}" for q in range(10)])
    sheet.append([])
    sheet.append(["1 - HFI"])
    fixed = [80e6] * 10
    variable = [2_696e6, 1_677e6] + [1_011e6] * 8
    hidden = [91.7e6, 58.1e6] + [36.1e6] * 8
    sheet.append([None, "x", "Fixed Income"] + fixed)
    sheet.append([None, "x", "Variable Rate Income"] + variable)
    sheet.append([None, "x", "Total"] + [f + v + h for f, v, h in zip(fixed, variable, hidden)])
    sheet.append([])
    sheet.append(["9, 10, 11"])
    sheet.append([None, "x", "Fixed Income"] + ["-"] * 10)
    merged_var = [1_260e6, 658e6] + [264e6] * 8
    sheet.append([None, "x", "Variable Rate Income"] + merged_var)
    sheet.append([None, "x", "Total"] + merged_var)
    book.save(path)
    return path


def test_reference_results_parse_blocks_markers_and_dashes(tmp_path):
    from scb_ppnr.ingestion.loans_loader import load_reference_results

    path = _results_workbook(tmp_path)
    results = load_reference_results(LoansSheetSpec(workbook=path, results_sheet="Results"))

    cni = results[("Commercial and industrial", "HFI")]
    assert cni["fixed"][0] == pytest.approx(80.0)            # dollars -> millions
    assert cni["variable"][1] == pytest.approx(1_677.0)
    # the hidden stream survives inside Total
    assert cni["total"][2] - cni["fixed"][2] - cni["variable"][2] == pytest.approx(36.1)

    merged = results[("Loans for Purchasing and Carrying Securities", "MERGED")]
    assert merged["fixed"][3] == 0.0                         # "-" means zero
    assert merged["total"][9] == pytest.approx(264.0)


def test_reference_results_requires_the_sheet_to_be_configured(tmp_path):
    from scb_ppnr.ingestion.loans_loader import load_reference_results

    path = _workbook(tmp_path, [_row("F1", 4, 1, 3)])
    with pytest.raises(ValidationFailure, match="results_sheet is not configured"):
        load_reference_results(LoansSheetSpec(workbook=path))


# --- outstanding balance (share basis, 2026-08-12) --------------------------

def test_outstanding_column_is_locom_dependent(tmp_path):
    """HFI rows read 'Launchpoint Outstanding Balance'; HFS/FVO rows read
    'Value' — the columns the reference workbook's share mix is built from."""
    headers = H1_HEADERS + ["Launchpoint Outstanding Balance", "Value"]
    rows = [
        _row("HFI-1", 4, 2, 3) + [750_000.0, None],       # HFI -> outstanding column
        _row("HFS-1", 4, 2, 1) + [None, 250_000.0],       # LOCOM 1 -> Value column
        _row("FVO-1", 4, 2, 2) + [None, 100_000.0],       # LOCOM 2 -> Value column
    ]
    path = _workbook(tmp_path, rows, h1_headers=headers)
    facilities, _ = load_facilities(LoansSheetSpec(workbook=path))

    assert facilities[0].outstanding_balance == pytest.approx(0.75)   # dollars -> millions
    assert facilities[1].outstanding_balance == pytest.approx(0.25)
    assert facilities[2].outstanding_balance == pytest.approx(0.10)


def test_outstanding_share_basis_changes_the_mix(tmp_path):
    """The root cause of the first-run gap: fixed/mixed books are fully drawn
    while revolvers and fee lines are not, so the committed mix underweights
    them. Outstanding basis restores the reference's mix."""
    from scb_ppnr.interest_income.loans_launchpoint import build_launch_point
    from scb_ppnr.interest_income.loans_schemas import SegmentKey

    headers = H1_HEADERS + ["Launchpoint Outstanding Balance", "Value"]
    rows = [
        # committed 100 each: committed shares 50/50 — but outstanding 90 vs 10
        _row("FIX", 4, 1, 3, rate=0.06, committed=100e6,
             originated="15-Jan-2020") + [90_000_000.0, None],
        _row("FLT", 4, 2, 3, rate=0.05, committed=100e6) + [10_000_000.0, None],
    ]
    path = _workbook(tmp_path, rows, h1_headers=headers)
    facilities, _ = load_facilities(LoansSheetSpec(workbook=path))
    history = {"2020Q1": 0.015, "2024Q4": 0.044}

    launch, _ = build_launch_point(
        facilities, {"Commercial and industrial": 1000.0}, 0.044, history,
        (1, 2, 3), lambda when: None, share_measure="outstanding",
    )
    fixed = launch[SegmentKey("Commercial and industrial", "HFI", 1)]
    floating = launch[SegmentKey("Commercial and industrial", "HFI", 2)]

    assert fixed.share == pytest.approx(0.9)
    assert fixed.balance == pytest.approx(900.0)
    assert floating.share == pytest.approx(0.1)
    # rate pools stay COMMITTED-weighted (they matched the reference exactly)
    assert floating.spread.pool_rate == pytest.approx(0.05)


def test_outstanding_basis_refuses_rows_without_the_column(tmp_path):
    from scb_ppnr.interest_income.loans_launchpoint import build_launch_point

    path = _workbook(tmp_path, [_row("F1", 4, 2, 3)])     # no outstanding columns
    facilities, _ = load_facilities(LoansSheetSpec(workbook=path))

    with pytest.raises(ValidationFailure, match="share_basis is 'outstanding'"):
        build_launch_point(
            facilities, {"Commercial and industrial": 1000.0}, 0.044, {}, (1, 2, 3),
            lambda when: None, share_measure="outstanding",
        )


def test_missing_outstanding_is_censused_with_row_and_id(tmp_path):
    """A [NULL] outstanding cell is counted and located BEFORE the launch-point
    layer refuses, so a ValidationFailure naming a Customer ID can be traced to
    its sheet row from the census alone."""
    headers = H1_HEADERS + ["Launchpoint Outstanding Balance", "Value"]
    rows = [
        _row("OK-1", 4, 2, 3) + [500_000.0, None],
        _row("0021575908", 4, 2, 3) + ["[NULL]", None],
    ]
    path = _workbook(tmp_path, rows, h1_headers=headers)
    _, census = load_facilities(LoansSheetSpec(workbook=path))

    assert census.missing_outstanding == 1
    assert census.missing_outstanding_rows == [(6, "0021575908")]
    assert "missing Outstanding/Value   : 1" in census.render()
    # both columns exist, so no configuration hint fires
    assert not any("share-basis column" in note for note in census.warnings)


def test_absent_share_columns_produce_a_configuration_hint(tmp_path):
    path = _workbook(tmp_path, [_row("F1", 4, 2, 3)])   # fixture lacks the columns
    _, census = load_facilities(LoansSheetSpec(workbook=path))

    assert census.missing_outstanding == 1
    assert any("col_outstanding / col_value" in note for note in census.warnings)


def test_m1_balances_split_by_side(tmp_path):
    """The reference blocks are per LOCOM side; each block's balance base is the
    M.1 balance of ITS side: first column of each dom/int pair = HFI at AC,
    second = HFS/FVO."""
    path = _m1_workbook(tmp_path, [
        _m1_row("Wholesale - Corp - C&I and others", "Wholesale - Corp - C&I and others",
                dom_hfi=135_038.0, dom_hfs=3_441.0, int_hfi=31_956.0, int_hfs=3_848.0),
    ])
    total, by_side, _ = load_category_balances(LoansSheetSpec(workbook=path))

    assert total["Commercial and industrial"] == pytest.approx(174_283.0)
    assert by_side[("Commercial and industrial", "HFI")] == pytest.approx(166_994.0)
    assert by_side[("Commercial and industrial", "FVO_HFS")] == pytest.approx(7_289.0)


def test_reference_engine_reproduces_the_workbook_construction(tmp_path):
    """engine='reference' (the user-supplied cell formulas, 2026-08-12):
    v3 merges into the variable segment at the FLOATING spread; balances are
    M.1(side) x outstanding-share within the side; the variable floor is
    outstanding-weighted with blank floors counting as ZERO; fixed floors at 0."""
    from scb_ppnr.interest_income.loans_launchpoint import build_launch_point
    from scb_ppnr.interest_income.loans_schemas import SegmentKey, VT_FIXED, VT_FLOATING

    headers = H1_HEADERS + ["Launchpoint Outstanding Balance", "Value"]
    rows = [
        # committed equal, outstanding very different -> shares must follow outstanding
        _row("FIX", 4, 1, 3, rate=0.055, committed=100e6,
             originated="15-Jan-2020") + [30_000_000.0, None],
        _row("FLT", 4, 2, 3, rate=0.060, committed=100e6, floor=0.04) + [50_000_000.0, None],
        _row("MIX", 4, 3, 3, rate=0.050, committed=100e6,
             originated="17-May-2022") + [20_000_000.0, None],
    ]
    path = _workbook(tmp_path, rows, h1_headers=headers)
    facilities, _ = load_facilities(LoansSheetSpec(workbook=path))
    side_balances = {("Commercial and industrial", "HFI"): 200.0}
    history = {"2020Q1": 0.015, "2022Q2": 0.008, "2024Q4": 0.044}

    launch, _ = build_launch_point(
        facilities, {}, 0.044, history, (1, 2, 3), lambda when: None,
        share_measure="outstanding", engine="reference", side_balances=side_balances,
    )

    variable = launch[SegmentKey("Commercial and industrial", "HFI", VT_FLOATING)]
    fixed = launch[SegmentKey("Commercial and industrial", "HFI", VT_FIXED)]

    # v3 merged into the variable segment: outstanding share (50+20)/100 of M.1(side)
    assert variable.share == pytest.approx(0.7)
    assert variable.balance == pytest.approx(140.0)
    # spread = FLOATING pool (committed-weighted over v2+v3): (0.060+0.050)/2 - 0.044
    assert variable.spread.pool_rate == pytest.approx(0.055)
    assert variable.spread.base_quarter is None
    # floor: out-weighted with the MIX row's blank floor counting as zero
    assert variable.floor == pytest.approx((50 * 0.04 + 20 * 0.0) / 70)
    # fixed: out-share 30/100 of the side balance, floored at exactly zero
    assert fixed.balance == pytest.approx(60.0)
    assert fixed.floor == 0.0
    # no separate mixed segment survives
    assert SegmentKey("Commercial and industrial", "HFI", 3) not in launch


def test_reference_merged_bucket_donor_pool_includes_mixed(tmp_path):
    """The reference's float-pool convention includes MIXED rows; under
    include_mixed the depository donor pool does too, pulling the borrowed
    rate up when mixed carries higher (past-set) rates."""
    from scb_ppnr.ingestion.loans_mapping import DEPOSITORY_INSTITUTION_H1_CODES
    from scb_ppnr.interest_income.loans_launchpoint import merged_bucket_launch_point

    path = _workbook(tmp_path, [
        _row("D-FLT", 1, 2, 3, rate=0.054, committed=800e6),
        _row("D-MIX", 1, 3, 3, rate=0.062, committed=200e6),
    ])
    facilities, _ = load_facilities(LoansSheetSpec(workbook=path))

    floating_only = merged_bucket_launch_point(
        facilities, 100.0, 0.044, DEPOSITORY_INSTITUTION_H1_CODES, "merged"
    )
    with_mixed = merged_bucket_launch_point(
        facilities, 100.0, 0.044, DEPOSITORY_INSTITUTION_H1_CODES, "merged", include_mixed=True
    )

    assert floating_only.spread.pool_rate == pytest.approx(0.054)
    assert with_mixed.spread.pool_rate == pytest.approx(0.8 * 0.054 + 0.2 * 0.062)
