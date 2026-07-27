"""Securities workbook loader (PID-SEC-6): the native multi-sheet workbook →
canonical `SecurityPosition`s grouped by model.

Contract: `specifications/interest-income/securities/securities-input-contract.md`.
Positions sheet is located by its MDRM header row (the row containing CQSCP084);
enrichment tabs are merged by CUSIP/ISIN; the prepayment pivot (PID-MBS-1) is
located by its "Row Labels" header with month-offset columns 0,3,…,27 ≙ PQ0..PQ9
(Grand Total ignored). Three date encodings are normalized here (yyyymmdd
integers; Excel serials; mm/dd/yyyy strings); scales are declared, never guessed
(D-006). Policy: unmapped categories and unmatched enrichment are hard errors
(PID-SEC-5 — surface and ask); equity-intent and out-of-scope positions are
excluded with a count; non-positive WAL rows are skipped with a highlighted
warning (user-parked 2026-07-24); PID-SEC-3 proxies amortized cost from
price/100 × current face when maturity/book-yield/AC are missing or zero.

PID-SEC-9 (2026-07-24): the positions sheet may carry the reference workbook's
own model-input columns on the MDRM header row (decimal units). When declared
in config they act as: maturity-years FALLBACK when the enrichment maturity
date is missing (feeds the PID-SEC-8 denominator — removes the AA-held-at-0
gap), coupon FALLBACK when the enrichment coupon is blank (rescues
FLOATER-NO-COUPON rows), and PREFERRED floor source over the enrichment floor
(drives floor_mode 'security_floor_else_zero')."""

from __future__ import annotations

import datetime as _dt
import math
from collections import Counter
from pathlib import Path

from ..core.schemas import PROJECTION_QUARTERS, ValidationFailure
from ..interest_income.securities_schemas import (
    MODEL_MBS,
    MODEL_OTHER_SEC,
    MODEL_UST,
    OUT_OF_SCOPE,
    RATE_FIXED,
    RATE_FLOATING,
    RATE_ZERO_COUPON,
    SecuritiesInputs,
    SecurityPosition,
    assign_model,
)
from .config import IngestionConfig, SecuritiesConfig, SecuritiesEnrichmentSheet
from .normalize import apply_money_scale, apply_rate_scale, to_float

_EXCEL_EPOCH = _dt.date(1899, 12, 30)
# Positions-sheet columns located via the MDRM header row (contract §2).
_POSITIONS_MDRM = {
    "D_DT": "report_date",
    "CQSCP083": "identifier_value",
    "CQSCP084": "security_description_1",
    "CQSCP087": "amortized_cost",
    "CQSCP089": "current_face_value",
    "CQSCP090": "original_face_value",
    "CQSCP092": "accounting_intent",
    "CQSCP094": "book_yield",
    "CQSCS383": "unique_id",             # per-row transaction id — positions sheets carry
                                         # multiple rows (lots) per CUSIP
}
# The price column is handled separately: configurable MDRM ([firm_data.securities].price_mdrm,
# default CQSCJH21) with a technical-name fallback ("Price") — see _positions_rows.
_REQUIRED_MDRM = ("CQSCP083", "CQSCP084", "CQSCP087", "CQSCP089", "CQSCP092")

_RATE_TYPE_MAP = {
    "FIXED": RATE_FIXED,
    "FLOATING": RATE_FLOATING,
    "FLOAT": RATE_FLOATING,
    "VARIABLE": RATE_FLOATING,           # company vocabulary (2026-07-24 diagnostics): ARM pools
    "ZERO COUPON": RATE_ZERO_COUPON,
    "ZERO_COUPON": RATE_ZERO_COUPON,
    "ZERO-COUPON": RATE_ZERO_COUPON,
}
# Step-coupon labels: the Fed model has no step machinery [FACT absence]. INTERIM
# treatment (pending user confirmation): fixed at the launch coupon, logged per security.
_STEP_LABELS = {"STEP CPN", "STEP", "STEP COUPON", "STEP-UP"}

# Excel formula-error literals (user-directed 2026-07-27): the reference sheet's
# derived columns (e.g. "Maturity (yr)", "Coupon Rate Floor (decimal)") are
# formulas that error to #VALUE! when their own inputs are missing — an error
# cell therefore means "no value", and the field falls back per its normal
# chain. A per-run census warning reports how many were seen, per field.
_EXCEL_ERRORS = {"#VALUE!", "#N/A", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}
_MISSING_STRINGS = {"", "(BLANK)", "N/A", "NA", "-"} | _EXCEL_ERRORS


def _load_workbook(path: Path):
    try:
        import openpyxl
    except ImportError:
        raise ValidationFailure(
            f"{path}: reading the securities workbook requires the optional 'openpyxl' "
            f"dependency (pip install openpyxl, or install with the [excel] extra)"
        ) from None
    if not path.exists():
        raise ValidationFailure(f"securities workbook not found: {path}")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _sheet(workbook, name: str, path: Path):
    if name not in workbook.sheetnames:
        raise ValidationFailure(f"{path}: sheet {name!r} not found; available: {workbook.sheetnames}")
    return workbook[name]


def _blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip().upper() in _MISSING_STRINGS)


def _technical_columns(sc: SecuritiesConfig) -> dict[str, str]:
    """PID-SEC-9 optional positions-sheet input columns: record field → exact header text."""
    declared = {
        "tech_maturity_years": sc.positions_maturity_years_column,
        "tech_coupon": sc.positions_coupon_column,
        "tech_floor": sc.positions_floor_column,
        "tech_rate_type": sc.positions_rate_type_column,
    }
    return {field: name for field, name in declared.items() if name is not None}


def _parse_date(value: object, *, context: str) -> _dt.date:
    """Normalize the contract's three date encodings (§6)."""
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if 10_000_000.0 <= number <= 99_991_231.0:       # yyyymmdd integer
            text = str(int(number))
            return _dt.date(int(text[:4]), int(text[4:6]), int(text[6:8]))
        if 0.0 < number < 200_000.0:                      # Excel serial
            return _EXCEL_EPOCH + _dt.timedelta(days=int(round(number)))
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try:
                return _dt.datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        try:
            return _parse_date(float(text.replace(",", "")), context=context)
        except (ValueError, ValidationFailure):
            pass
    raise ValidationFailure(f"{context}: cannot interpret {value!r} as a date (yyyymmdd, Excel serial, or mm/dd/yyyy)")


def _positions_rows(
    worksheet, path: Path, price_mdrm: str = "CQSCJH21",
    technical_columns: dict[str, str] | None = None,
) -> tuple[list[dict[str, object]], dict[str, int], list[str]]:
    rows = list(worksheet.iter_rows(values_only=True))
    notes: list[str] = []
    header_index: int | None = None
    columns: dict[str, int] = {}
    for index, row in enumerate(rows):
        cells = {str(c).strip(): i for i, c in enumerate(row) if c is not None and str(c).strip()}
        if "CQSCP084" in cells:
            header_index = index
            for mdrm, field in _POSITIONS_MDRM.items():
                if mdrm in cells:
                    columns[field] = cells[mdrm]
            if price_mdrm in cells:
                columns["price"] = cells[price_mdrm]
            # PID-SEC-9: config-declared technical input columns on the same header row.
            for field, name in (technical_columns or {}).items():
                if name in cells:
                    columns[field] = cells[name]
                else:
                    raise ValidationFailure(
                        f"{path}: configured positions-sheet column {name!r} (PID-SEC-9) not found "
                        f"on the MDRM header row — check the exact header text in the config"
                    )
            # Reference per-security income columns (verification only): the workbook's
            # own II_PQ1..II_PQ9, when present on the same header row.
            reference_columns = {f"II_PQ{q}": cells[f"II_PQ{q}"] for q in range(1, 10) if f"II_PQ{q}" in cells}
            if len(reference_columns) == 9:
                for q in range(1, 10):
                    columns[f"reference_pq{q}"] = reference_columns[f"II_PQ{q}"]
                notes.append("reference income columns II_PQ1..II_PQ9 found — attached for compare mode")
            break
    if header_index is None:
        raise ValidationFailure(f"{path}: positions sheet has no MDRM header row (looked for CQSCP084)")
    if "price" not in columns:
        # Fallback: locate a header cell (any of the header rows above the data) whose
        # technical name is exactly "Price" (case-insensitive).
        for row in rows[: header_index + 1]:
            for i, cell in enumerate(row):
                if isinstance(cell, str) and cell.strip().lower() == "price":
                    columns["price"] = i
                    notes.append(
                        f"price column located by technical name 'Price' (column {i + 1}) — the "
                        f"configured MDRM {price_mdrm!r} was not found; set "
                        f"[firm_data.securities].price_mdrm if the workbook uses a different code"
                    )
                    break
            if "price" in columns:
                break
    if "price" not in columns:
        notes.append(
            f"NO price column found (MDRM {price_mdrm!r} absent and no 'Price' technical name) — "
            f"PID-SEC-3 proxies are impossible; securities with missing/zero amortized cost will stop the run"
        )
    missing = [m for m in _REQUIRED_MDRM if _POSITIONS_MDRM[m] not in columns]
    if missing:
        raise ValidationFailure(f"{path}: positions MDRM header lacks required columns {missing}")

    parsed: list[dict[str, object]] = []
    for row in rows[header_index + 1:]:
        record = {field: (row[i] if i < len(row) else None) for field, i in columns.items()}
        if _blank(record.get("identifier_value")):
            continue
        parsed.append(record)
    return parsed, columns, notes


def _enrichment_map(workbook, spec: SecuritiesEnrichmentSheet, path: Path) -> dict[str, dict[str, object]]:
    worksheet = _sheet(workbook, spec.sheet, path)
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < spec.header_row:
        raise ValidationFailure(f"{path}:{spec.sheet}: header_row {spec.header_row} beyond sheet end")
    header = {str(c).strip(): i for i, c in enumerate(rows[spec.header_row - 1]) if c is not None and str(c).strip()}
    needed = {"key": spec.key_column, "maturity": spec.maturity_column, "coupon": spec.coupon_column,
              "rate_type": spec.rate_type_column, "wal": spec.wal_column}
    if spec.floor_column is not None:
        needed["floor"] = spec.floor_column
    if spec.floater_indicator_column is not None:
        needed["floater_indicator"] = spec.floater_indicator_column
    columns: dict[str, int] = {}
    for role, name in needed.items():
        if name not in header:
            raise ValidationFailure(
                f"{path}:{spec.sheet}: column {name!r} ({role}) not found in header row "
                f"{spec.header_row}; available: {sorted(header)}"
            )
        columns[role] = header[name]

    result: dict[str, dict[str, object]] = {}
    for row in rows[spec.header_row:]:
        key_raw = row[columns["key"]] if columns["key"] < len(row) else None
        if _blank(key_raw):
            continue
        key = str(key_raw).strip()
        result[key] = {role: (row[i] if i < len(row) else None) for role, i in columns.items()}
    return result


def _prepayment_map(worksheet, path: Path, money_scale: str, warnings: list[str]) -> dict[str, dict[int, float]]:
    rows = list(worksheet.iter_rows(values_only=True))
    header_index = None
    for index, row in enumerate(rows):
        first = row[0] if row else None
        if isinstance(first, str) and first.strip().lower() == "row labels":
            header_index = index
            break
    if header_index is None:
        raise ValidationFailure(f"{path}: prepayment sheet has no 'Row Labels' header row (pivot layout, contract §4)")
    month_columns: dict[int, int] = {}
    for i, cell in enumerate(rows[header_index][1:], start=1):
        if cell is None or (isinstance(cell, str) and not cell.strip().lstrip("-").isdigit()):
            continue
        months = int(float(cell)) if not isinstance(cell, int) else cell
        if months % 3 == 0 and 0 <= months <= 27:
            month_columns[months // 3] = i
    missing_quarters = [q for q in (0, *PROJECTION_QUARTERS) if q not in month_columns]
    if missing_quarters:
        raise ValidationFailure(
            f"{path}: prepayment header lacks month-offset columns for PQ{missing_quarters} "
            f"(expected 0,3,…,27)"
        )

    faces: dict[str, dict[int, float]] = {}
    for line, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        label = row[0] if row else None
        if _blank(label) or str(label).strip().lower() == "grand total":
            continue
        cusip = str(label).strip()
        # Pivot semantics (user-directed 2026-07-24): blank cells ARE 0 — no warning,
        # matching the reference workbook's own treatment.
        path_values: dict[int, float] = {}
        for quarter, column in month_columns.items():
            raw = row[column] if column < len(row) else None
            if _blank(raw):
                path_values[quarter] = 0.0
            else:
                value = to_float(raw, context=f"{path} prepayment {cusip} PQ{quarter}")
                path_values[quarter] = apply_money_scale(money_scale, value, context=f"{path} prepayment {cusip} PQ{quarter}")
        # User-directed rule (2026-07-24): a row whose PQ1 face is 0 (or blank) is skipped at load.
        if path_values[1] == 0.0:
            warnings.append(
                f"prepayment row {cusip}: PQ1 face is 0 — row skipped (user-directed 2026-07-24); "
                f"if the security is still an in-scope Agency MBS position it falls back to the "
                f"flat-face (no-prepayment) treatment"
            )
            continue
        faces[cusip] = path_values
    return faces


def load_securities_inputs(config: IngestionConfig) -> SecuritiesInputs:
    if config.firm_data is None or config.firm_data.securities is None:
        raise ValidationFailure("config has no [firm_data.securities] section (PID-SEC-6)")
    sc: SecuritiesConfig = config.firm_data.securities
    path = config.resolve(sc.workbook)
    workbook = _load_workbook(path)
    warnings: list[str] = []
    try:
        records, _, header_notes = _positions_rows(
            _sheet(workbook, sc.positions_sheet, path), path, sc.price_mdrm, _technical_columns(sc)
        )
        if not records:
            raise ValidationFailure(f"{path}: positions sheet has no data rows")
        enrichment: dict[str, dict[str, object]] = {}
        for spec in sc.enrichment:
            for key, fields in _enrichment_map(workbook, spec, path).items():
                enrichment.setdefault(key, fields)      # first tab wins on duplicates
        prepayment: dict[str, dict[int, float]] = {}
        if sc.prepayment_sheet is not None:
            prepayment = _prepayment_map(_sheet(workbook, sc.prepayment_sheet, path), path, sc.money_scale, warnings)
    finally:
        workbook.close()
    warnings.extend(header_notes)

    report_date = _parse_date(records[0]["report_date"], context=f"{path} D_DT") if "report_date" in records[0] and not _blank(records[0].get("report_date")) else None

    # Excel error-cell census (user-directed 2026-07-27): counted per field so the
    # treated-as-missing rule stays surfaced, never silent.
    error_cells: Counter[str] = Counter()
    for record in records:
        for field, value in record.items():
            if isinstance(value, str) and value.strip().upper() in _EXCEL_ERRORS:
                error_cells[field] += 1
    for fields in enrichment.values():
        for role, value in fields.items():
            if isinstance(value, str) and value.strip().upper() in _EXCEL_ERRORS:
                error_cells[f"enrichment.{role}"] += 1
    if error_cells:
        breakdown = ", ".join(f"{field}×{n}" for field, n in sorted(error_cells.items()))
        warnings.append(
            f"{sum(error_cells.values())} Excel error cell(s) (#VALUE! etc.) treated as missing "
            f"(user-directed 2026-07-27); each field falls back per its normal chain: {breakdown}"
        )

    skipped_out_of_scope = 0
    skipped_wal = 0
    tech_maturity_rows = 0
    tech_coupon_rows = 0
    tech_floor_rows = 0
    indicator_disagree = 0
    unmatched: list[str] = []
    grouped: dict[str, list[SecurityPosition]] = {MODEL_UST: [], MODEL_MBS: [], MODEL_OTHER_SEC: []}

    # Multi-row CUSIPs (lots): the prepayment path is applied PER ROW as a survival
    # factor scaled by each lot's own launch face (user point 6, 2026-07-24).
    total_rows_per_cusip: dict[str, int] = {}
    for record in records:
        cusip = str(record["identifier_value"]).strip()
        total_rows_per_cusip[cusip] = total_rows_per_cusip.get(cusip, 0) + 1
    row_counter_per_cusip: dict[str, int] = {}
    used_prepayment: set[str] = set()
    apportioned_cusips: set[str] = set()

    for record in records:
        cusip = str(record["identifier_value"]).strip()
        row_counter_per_cusip[cusip] = row_counter_per_cusip.get(cusip, 0) + 1
        unique = "" if _blank(record.get("unique_id")) else str(record["unique_id"]).strip()
        if unique:
            security_id = unique
        elif total_rows_per_cusip.get(cusip, 1) == 1:
            security_id = cusip                              # single lot — keep the plain CUSIP
        else:
            security_id = f"{cusip}#r{row_counter_per_cusip[cusip]}"
        context = f"{path} security {cusip} ({security_id})"
        intent = "" if _blank(record.get("accounting_intent")) else str(record["accounting_intent"]).strip()
        category = str(record["security_description_1"]).strip() if not _blank(record.get("security_description_1")) else ""
        if not category:
            raise ValidationFailure(f"{context}: security_description_1 is blank")
        model, agency_prepay = assign_model(category)
        if model == OUT_OF_SCOPE or intent.upper() == "EQ":
            skipped_out_of_scope += 1
            continue

        fields = enrichment.get(cusip)
        if fields is None:
            # User-visible skip (2026-07-24, pending confirmation): previously a hard stop.
            unmatched.append(cusip)
            warnings.append(f"HIGHLIGHT {cusip}: no enrichment match — position skipped (surfaced; confirm skip vs data fix)")
            continue

        rate_type_raw = "" if _blank(fields.get("rate_type")) else str(fields["rate_type"]).strip().upper()
        if rate_type_raw in _STEP_LABELS:
            rate_type = RATE_FIXED
            warnings.append(
                f"{security_id}: step-coupon rate type {rate_type_raw!r} treated as FIXED at the "
                f"launch coupon (confirmed vs the reference 2026-07-24: coupon held flat)"
            )
        elif rate_type_raw in _RATE_TYPE_MAP:
            rate_type = _RATE_TYPE_MAP[rate_type_raw]
        else:
            raise ValidationFailure(f"{context}: unknown rate type {fields.get('rate_type')!r} (known: {sorted(_RATE_TYPE_MAP)} + step labels {sorted(_STEP_LABELS)})")

        # Optional isFloaterIndicator cross-check (consistency monitor — logs, never blocks):
        # the flag is redundant with rate_type, which is what disagreement makes it useful for.
        if not _blank(fields.get("floater_indicator")):
            flag_raw = str(fields["floater_indicator"]).strip().upper()
            if flag_raw in ("Y", "YES", "TRUE", "1"):
                flag = True
            elif flag_raw in ("N", "NO", "FALSE", "0"):
                flag = False
            else:
                flag = None
                warnings.append(f"{security_id}: floater-indicator value {fields['floater_indicator']!r} not recognized (monitor only)")
            if flag is not None and flag != (rate_type == RATE_FLOATING):
                warnings.append(
                    f"{security_id}: floater-indicator {flag_raw!r} disagrees with rate type "
                    f"{rate_type_raw!r} — data-quality monitor, rate type governs (surfaced, not blocked)"
                )

        # PID-SEC-9: the positions-sheet coupon (decimal) is always carried for
        # source-difference diagnostics; it drives the model only as a blank-fill.
        excel_coupon = None
        if "tech_coupon" in record and not _blank(record.get("tech_coupon")):
            excel_coupon = to_float(record["tech_coupon"], context=f"{context} coupon (positions sheet, decimal)")
        coupon = None
        if not _blank(fields.get("coupon")):
            coupon = apply_rate_scale(sc.coupon_scale, to_float(fields["coupon"], context=f"{context} coupon"), context=f"{context} coupon")
        elif excel_coupon is not None:
            coupon = excel_coupon
            tech_coupon_rows += 1
            warnings.append(f"{security_id}: enrichment coupon blank — positions-sheet coupon column used (decimal; PID-SEC-9)")
        if rate_type == RATE_ZERO_COUPON:
            coupon = None                                   # zero-coupon: cash coupon is identically zero
        floor = None
        if "tech_floor" in record and not _blank(record.get("tech_floor")):
            # PID-SEC-9: the positions-sheet floor column (decimal) is PREFERRED —
            # it is the reference workbook's own floor input.
            floor = to_float(record["tech_floor"], context=f"{context} floor (positions sheet, decimal)")
            tech_floor_rows += 1
        elif "floor" in (fields or {}) and not _blank(fields.get("floor")):
            floor = apply_rate_scale(sc.coupon_scale, to_float(fields["floor"], context=f"{context} floor"), context=f"{context} floor")
        # PID-SEC-9: the sheet's own float/fixed indicator — carried verification-only
        # (bake-off + monitor); the ITO rate type still drives the models.
        excel_rate_label = None
        if "tech_rate_type" in record and not _blank(record.get("tech_rate_type")):
            excel_rate_label = str(record["tech_rate_type"]).strip()
            upper = excel_rate_label.upper()
            if ("FIX" in upper and rate_type == RATE_FLOATING) or ("FLOAT" in upper and rate_type == RATE_FIXED):
                indicator_disagree += 1
        # Data-quality monitor (2026-07-27, motivated by 300–470% OURS-HIGH CLO floaters
        # under floor_mode 'security_floor_else_zero'): a floor far above the coupon —
        # or above 25% annualized — is almost certainly a mis-unit source cell. Logged
        # and applied AS-IS (log-never-clamp); the RATE_SCALE_GUARD (50%) still hard-stops.
        if floor is not None and (floor > 0.25 or (coupon is not None and floor > 3.0 * max(coupon, 0.02))):
            warnings.append(
                f"HIGHLIGHT {security_id}: coupon floor {floor} looks implausible vs coupon {coupon} — "
                f"suspect source-cell units (applied AS-IS, log-never-clamp; confirm the cell)"
            )

        wal = None
        if not _blank(fields.get("wal")):
            wal = to_float(fields["wal"], context=f"{context} wal")
            if wal <= 0.0:
                skipped_wal += 1
                warnings.append(
                    f"HIGHLIGHT {security_id}: non-positive WAL {wal} — row skipped "
                    f"(user-parked 2026-07-24; treatment to be decided)"
                )
                continue

        # Two maturity-year candidates (PID-SEC-8 denominator + event timing): the
        # ITO maturity date ÷ 365 [PID-SEC-8, user-verified], and the sheet's own
        # years column [PID-SEC-9]. `maturity_source` (OQ-030) picks which is
        # primary — "positions_first" replicates the reference's actual divisor
        # (for callable munis the sheet column can be the call-adjusted maturity,
        # shorter than the ITO final-maturity date).
        enrich_years = None
        if not _blank(fields.get("maturity")) and report_date is not None:
            maturity_date = _parse_date(fields["maturity"], context=f"{context} maturity")
            days = (maturity_date - report_date).days
            if days <= 0:
                warnings.append(f"{security_id}: maturity {maturity_date} not after the report date {report_date} — treated as missing (PID-SEC-3 may proxy)")
            else:
                enrich_years = days / 365.0
        tech_years = None
        if "tech_maturity_years" in record and not _blank(record.get("tech_maturity_years")):
            candidate = to_float(record["tech_maturity_years"], context=f"{context} maturity years (positions sheet)")
            if candidate > 0.0:
                tech_years = candidate
            else:
                warnings.append(f"{security_id}: positions-sheet maturity years {candidate} not positive — ignored (surfaced)")
        if sc.maturity_source == "positions_first":
            maturity_years = tech_years if tech_years is not None else enrich_years
            used_tech = tech_years is not None
        else:
            maturity_years = enrich_years if enrich_years is not None else tech_years
            used_tech = enrich_years is None and tech_years is not None
        maturity_quarters = max(1, math.ceil(4.0 * maturity_years)) if maturity_years is not None else None
        if used_tech:
            tech_maturity_rows += 1
        # PID-SEC-7 (user-confirmed 2026-07-24): Agency MBS without a maturity date
        # use WAL as the maturity in years; quarters = ceil(4 × WAL).
        if maturity_years is None and agency_prepay and wal is not None:
            maturity_years = wal
            maturity_quarters = max(1, math.ceil(4.0 * wal))
            warnings.append(
                f"{security_id}: Agency MBS without a maturity date — WAL {wal} years used as "
                f"maturity ({maturity_quarters} quarters; company convention, PID-SEC-7)"
            )

        face = apply_money_scale(sc.money_scale, to_float(record["current_face_value"], context=f"{context} current_face"), context=f"{context} current_face")
        ac_raw = None if _blank(record.get("amortized_cost")) else to_float(record["amortized_cost"], context=f"{context} amortized_cost")
        amortized_cost = None if ac_raw is None else apply_money_scale(sc.money_scale, ac_raw, context=f"{context} amortized_cost")
        book_yield = None
        if not _blank(record.get("book_yield")):
            by = to_float(record["book_yield"], context=f"{context} book_yield")
            book_yield = apply_rate_scale(sc.book_yield_scale, by, context=f"{context} book_yield") if by != 0.0 else 0.0

        # PID-SEC-3: unsettled-transaction AC proxy.
        ac_proxied = False
        trigger = maturity_quarters is None or book_yield in (None, 0.0) or amortized_cost in (None, 0.0)
        if trigger and amortized_cost in (None, 0.0):
            if _blank(record.get("price")):
                raise ValidationFailure(
                    f"{context}: PID-SEC-3 trigger (maturity/book yield/amortized cost missing or zero) "
                    f"but no price to proxy from — surfaced, never defaulted"
                )
            price = to_float(record["price"], context=f"{context} price")
            amortized_cost = price / 100.0 * face          # per-100 price × notional (current face, TBC)
            ac_proxied = True
            warnings.append(f"{security_id}: PID-SEC-3 amortized-cost proxy applied (price/100 × current face); accretion held at 0")
        if book_yield == 0.0:
            book_yield = None

        reference_income = None
        if all(f"reference_pq{q}" in record for q in range(1, 10)):
            try:
                reference_income = {
                    q: apply_money_scale(sc.money_scale, to_float(record[f"reference_pq{q}"], context=f"{context} II_PQ{q}"),
                                         context=f"{context} II_PQ{q}")
                    for q in range(1, 10)
                    if not _blank(record.get(f"reference_pq{q}"))
                }
                if len(reference_income) != 9:
                    reference_income = None               # partial rows carry no reference
            except ValidationFailure:
                reference_income = None                   # unparseable reference cells — verification only

        face_path = None
        if agency_prepay and cusip in prepayment:
            used_prepayment.add(cusip)
            cusip_path = prepayment[cusip]
            base = cusip_path[0]
            if base <= 0.0:
                warnings.append(f"{security_id}: prepayment PQ0 face is 0 — face held flat (surfaced)")
            else:
                # PID-SEC-8 / user point 6 (2026-07-24): computed per row — the CUSIP path is
                # applied as a survival FACTOR scaled by this row's own launch face.
                face_path = {q: face * (value / base) for q, value in cusip_path.items()}
                if total_rows_per_cusip.get(cusip, 1) > 1:
                    apportioned_cusips.add(cusip)
        elif agency_prepay:
            warnings.append(f"{security_id}: Agency MBS absent from the prepayment sheet — no prepayment, face held flat (multi-family, PID-SEC-5)")

        grouped[model].append(
            SecurityPosition(
                security_id=security_id,
                model=model,
                category=category,
                rate_type=rate_type,
                accounting_intent=intent or "AFS",
                current_face=face,
                amortized_cost=amortized_cost if amortized_cost is not None else 0.0,
                coupon_rate=coupon,
                book_yield=book_yield,
                coupon_floor=floor,
                maturity_quarters=maturity_quarters,
                maturity_years=maturity_years,
                wal_years=wal,
                face_path=face_path,
                ac_proxied=ac_proxied,
                reference_income=reference_income,
                excel_rate_label=excel_rate_label,
                excel_coupon_rate=excel_coupon,
            )
        )

    if unmatched:
        warnings.append(
            f"{len(unmatched)} position(s) skipped for missing enrichment (maturity/coupon/rate-type "
            f"unavailable) — HIGHLIGHTED above; confirm skip vs data fix"
        )
    if apportioned_cusips:
        warnings.append(
            f"{len(apportioned_cusips)} multi-lot Agency CUSIP(s): prepayment applied per row as a "
            f"survival factor on each lot's own launch face (user point 6, 2026-07-24)"
        )
    for leftover in sorted(set(prepayment) - used_prepayment):
        warnings.append(f"prepayment sheet row {leftover} matches no in-scope Agency MBS position (ignored, logged)")
    if skipped_out_of_scope:
        warnings.append(f"{skipped_out_of_scope} equity-intent/out-of-scope position(s) excluded (PID-SEC-5)")
    if skipped_wal:
        warnings.append(f"{skipped_wal} position(s) skipped for non-positive WAL — HIGHLIGHTED for later decision")
    if tech_maturity_rows:
        warnings.append(
            f"{tech_maturity_rows} position(s): maturity years taken from the positions-sheet column "
            f"(maturity_source={sc.maturity_source!r}; PID-SEC-9/OQ-030) for the PID-SEC-8 denominator "
            f"and event timing"
        )
    if tech_floor_rows:
        warnings.append(f"{tech_floor_rows} position(s): floor taken from the positions-sheet floor column (decimal; PID-SEC-9)")
    if tech_coupon_rows:
        warnings.append(f"{tech_coupon_rows} position(s): coupon taken from the positions-sheet coupon column (decimal; PID-SEC-9)")
    if indicator_disagree:
        warnings.append(
            f"{indicator_disagree} position(s): positions-sheet float/fixed indicator disagrees with "
            f"the ITO rate type — ITO governs (verification-only monitor, PID-SEC-9)"
        )

    return SecuritiesInputs(
        firm_id=config.firm_data.firm_id,
        ust=tuple(grouped[MODEL_UST]),
        mbs=tuple(grouped[MODEL_MBS]),
        other_sec=tuple(grouped[MODEL_OTHER_SEC]),
        warnings=tuple(warnings),
    )
