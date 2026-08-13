"""Workbook binding for the Corporate wholesale loan model (PID-LOAN-8/9/10).

Reads the firm's single workbook and produces canonical objects the model layers
already consume: `LoanFacility` rows, the merged 9/10/11 bucket balance, and the
3-month Treasury history and projection. Every decision this file makes was
settled before it was written — decoding lives in `loans_mapping`, arithmetic in
`interest_income.loans_launchpoint` / `loans_projection`. This module only binds
physical cells to those.

Sheet names, header rows and column headers are configuration; only the logical
contract is committed here. The workbook itself never enters this repository.

Scales are declared and never guessed (D-006), which matters more than usual
because four different ones live in this one file: H.1 rates are decimals, H.1
exposures are whole dollars, the FR Y-9C extract is in thousands, and the MEV
rates are percentages."""

from __future__ import annotations

import datetime as _dt
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from types import MappingProxyType
from typing import Iterable, Mapping, Sequence

from ..core.schemas import ValidationFailure
from ..interest_income.loans_schemas import LoanFacility, check_quarter_label
from .loans_mapping import (
    FED_CATEGORY_NAMES,
    decode_segment,
    m1_role_category,
    parse_h1_code,
    parse_locom,
    reference_key,
)
from .normalize import (
    SCALE_DOLLARS,
    SCALE_PERCENT,
    SCALE_THOUSANDS,
    apply_money_scale,
    apply_rate_scale,
    to_float,
)

# Values that all mean "absent". `[NULL]` is the workbook's own token; the Excel
# formula errors follow the PID-SEC-6 amendment — a derived cell that errors
# because its own inputs are missing is a missing value, not a parse failure.
_EXCEL_ERRORS = {"#VALUE!", "#N/A", "#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}
_MISSING_TOKENS = {"", "NA", "N/A", "NONE", "[NULL]", "NULL", "-"} | _EXCEL_ERRORS

# The merged-bucket MDRMs (PID-LOAN-10).
MDRM_PURCHASING_CARRYING_SECURITIES = "BHCK1545"
MDRM_FARMLAND = "BHDM1420"
MERGED_BUCKET_MDRMS = (MDRM_PURCHASING_CARRYING_SECURITIES, MDRM_FARMLAND)

# The sheet spells this correctly (user-confirmed 2026-08-07 — the misspelling
# was in the message, not the workbook). The variant is still accepted second,
# because an extract that does carry it is not worth a failed run, and any
# substitution is reported rather than applied silently.
_VARIABILITY_ALIASES = ("Interest Rate Variability", "Interest Rate Variablility")


@dataclass(frozen=True)
class LoansSheetSpec:
    """Where the Corporate inputs live. Every field is company-local config."""

    workbook: Path
    h1_sheet: str = "CORP H.1"
    h1_header_row: int = 4
    fry9c_sheet: str = "FR-Y9C 4Q 2024"
    fry9c_header_row: int = 8
    fry9c_id_column: str = "ID_RSSD"
    fry9c_value_column_index: int = 3       # 1-based; "the third column"
    m1_sheet: str = "M.1 Balance"
    m1_first_data_row: int = 11
    m1_domestic_role_column_index: int = 1    # 1-based; column A carries the domestic role
    m1_international_role_column_index: int = 2               # column B, the international role
    m1_domestic_value_column_indices: tuple[int, ...] = (5, 7)    # E = HFI at AC, G = HFS/FVO
    m1_international_value_column_indices: tuple[int, ...] = (9, 11)   # I = HFI at AC, K = HFS/FVO
    m1_scale: str = "millions"
    mev_sheet: str = "MEV"
    mev_header_row: int = 1
    mev_scenario_column: str = "Scenario Name"
    mev_date_column: str = "Date"
    mev_3m_column: str = "3-month Treasury rate"
    mev_history_scenario: str = "Actual"
    # declared scales — the loader refuses to run without them
    rate_scale: str = "decimal"
    exposure_scale: str = SCALE_DOLLARS
    fry9c_scale: str = SCALE_THOUSANDS
    mev_rate_scale: str = SCALE_PERCENT
    # H.1 column headers
    col_facility_id: str = "Customer ID"
    col_h1_code: str = "Line Reported on FR Y9C"
    col_variability: str | None = None      # None -> try the aliases
    col_locom: str = "Lower of Cost or Market Flag"
    col_interest_rate: str = "Interest Rate"
    col_committed: str = "Committed Exposure Global"
    col_utilized: str = "Utilized Exposure Global"
    col_floor: str = "Interest Rate Floor"
    col_origination: str = "Origination Date"
    col_maturity: str = "Maturity Date"
    # Optional fallback identifier columns (PID-LOAN-12). A row whose Customer ID
    # is [NULL] tries these in order; a row where all three are [NULL] gets a
    # synthesized row-number label — its balances are real and must not be dropped.
    col_internal_id: str = "Internal ID"
    col_original_internal_id: str = "Original Internal ID"
    # Share-basis columns (2026-08-12): the reference workbook's segment balances
    # are OUTSTANDING-mix based — "Launchpoint Outstanding Balance" on HFI rows,
    # "Value" on HFS/FVO rows. Optional: looked up if present; required only when
    # share_basis = "outstanding".
    col_outstanding: str = "Launchpoint Outstanding Balance"
    col_value: str = "Value"
    # Reference results sheet (compare mode). The workbook's own projected income,
    # laid out as blocks per "N - HFI" / "N - HFS/FVO" marker (plus a "9, 10, 11"
    # merged block) with Fixed Income / Variable Rate Income / Total rows over
    # PQ0..PQ9 columns, in raw dollars. Optional: unset disables the compare.
    results_sheet: str | None = None
    # --- CRE part (PID-LOAN-18/19/20) ---------------------------------------
    # Setting cre_h2_sheet enables the CRE run; None keeps the run Corporate-only.
    # Column names default to the H.2 sheet's own headers (PID-LOAN-18); the
    # loader reuses the H.1 rate/floor/date/LOCOM headers where they are the
    # same, and the CRE-specific balance columns below where they differ.
    cre_h2_sheet: str | None = None
    cre_h2_header_row: int = 4
    cre_col_committed: str = "Committed Balance"
    cre_col_outstanding: str = "Outstanding Balance"
    cre_col_value: str | None = None       # optional distinct HFS/FVO weighting column
                                           # ("Launchpoint Value" residual, CRE brief §0.1 ii);
                                           # unset -> Outstanding Balance weights both sides
    # M.1 wiring is by ROW (PID-LOAN-20: the user confirmed the cells directly):
    # E/G of each row = the domestic HFI / HFS-FVO balances of its category;
    # the international category sums I/K over the same three rows. The value
    # column indices are shared with the Corporate M.1 read (E, G, I, K).
    cre_m1_construction_row: int = 17
    cre_m1_multifamily_row: int = 18
    cre_m1_non_owner_occupied_row: int = 21
    # The workbook's own CRE projected income (same block layout as Corporate's,
    # markers "1 - HFI" .. "4 - HFS/FVO", no merged block). Unset disables.
    cre_results_sheet: str | None = None


@dataclass
class LoaderCensus:
    """What the load found, reported rather than assumed.

    Missing rates and floors are expected and benign; a large count is not, so
    the exposure behind each is carried alongside the count."""

    rows_read: int = 0
    missing_interest_rate: int = 0
    missing_interest_rate_exposure: float = 0.0
    missing_floor: int = 0
    missing_origination_date: int = 0
    missing_maturity_date: int = 0
    missing_outstanding: int = 0
    missing_outstanding_rows: list = field(default_factory=list)   # (sheet row, facility id)
    id_sources: Counter = field(default_factory=Counter)
    unidentified_rows: list[int] = field(default_factory=list)
    unidentified_exposure: float = 0.0
    reference_keys: Counter = field(default_factory=Counter)
    column_substitutions: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    title: str = "LOANS LOADER CENSUS"
    # CRE-only counters (PID-LOAN-19): rows whose line code is DO NOT USE fall
    # outside every CRE category — excluded WITH their exposure on record, and
    # a blank Outstanding Balance is a genuine zero (an undrawn facility), not
    # a refused row.
    excluded_line_codes: Counter = field(default_factory=Counter)
    excluded_line_code_exposure: float = 0.0
    blank_outstanding: int = 0

    def render(self) -> str:
        lines = [self.title]
        lines.append(f"  rows read                   : {self.rows_read}")
        lines.append(
            f"  missing Interest Rate       : {self.missing_interest_rate}"
            f"  (exposure {self.missing_interest_rate_exposure:,.2f})"
        )
        lines.append(f"  no floor on file            : {self.missing_floor}")
        lines.append(f"  missing Origination Date    : {self.missing_origination_date}")
        lines.append(f"  missing Maturity Date       : {self.missing_maturity_date}")
        if self.missing_outstanding:
            shown = ", ".join(f"row {r} ({fid})" for r, fid in self.missing_outstanding_rows[:8])
            more = (f", ... {self.missing_outstanding - 8} more"
                    if self.missing_outstanding > 8 else "")
            lines.append(
                f"  missing Outstanding/Value   : {self.missing_outstanding}"
                f"  ({shown}{more}) — share_basis='outstanding' refuses these rows"
            )
        if self.id_sources:
            lines.append(
                "  facility-ID sources         : "
                + ", ".join(f"{name} {count}" for name, count in sorted(self.id_sources.items()))
            )
        if self.unidentified_rows:
            shown = ", ".join(str(r) for r in self.unidentified_rows[:15])
            more = f", ... {len(self.unidentified_rows) - 15} more" if len(self.unidentified_rows) > 15 else ""
            lines.append(
                f"  unidentified rows (labeled) : {len(self.unidentified_rows)}"
                f"  (committed exposure {self.unidentified_exposure:,.2f})  sheet rows: {shown}{more}"
            )
        if self.excluded_line_codes:
            detail = ", ".join(f"code {code} x{count}" for code, count in sorted(self.excluded_line_codes.items()))
            lines.append(
                f"  excluded DO-NOT-USE line-code rows : "
                f"{sum(self.excluded_line_codes.values())}  ({detail}; committed exposure "
                f"{self.excluded_line_code_exposure:,.2f}) — outside every CRE category "
                f"(PID-LOAN-19), never allocated balance"
            )
        if self.blank_outstanding:
            lines.append(
                f"  blank Outstanding Balance   : {self.blank_outstanding} rows read as 0 "
                f"(an undrawn facility has no outstanding balance — a genuine zero, not a refusal)"
            )
        lines.append(f"  distinct reference keys     : {len(self.reference_keys)}")
        for key, count in sorted(self.reference_keys.items()):
            lines.append(f"      {key:<14} {count}")
        for note in self.column_substitutions + self.warnings:
            lines.append(f"  NOTE: {note}")
        return "\n".join(lines)


def _open(path: Path):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment guard
        raise ValidationFailure("reading the loans workbook requires openpyxl") from exc
    if not path.exists():
        raise ValidationFailure(f"loans workbook not found: {path}")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _sheet(workbook, name: str, path: Path):
    if name not in workbook.sheetnames:
        raise ValidationFailure(
            f"{path}: sheet {name!r} not found; available: {sorted(workbook.sheetnames)}"
        )
    return workbook[name]


def _header_index(rows: Sequence[Sequence[object]], header_row: int, context: str) -> dict[str, int]:
    if len(rows) < header_row:
        raise ValidationFailure(f"{context}: header row {header_row} is beyond the end of the sheet")
    return {
        str(cell).strip(): index
        for index, cell in enumerate(rows[header_row - 1])
        if cell is not None and str(cell).strip()
    }


def _column(header: Mapping[str, int], name: str, context: str) -> int:
    if name not in header:
        raise ValidationFailure(
            f"{context}: column {name!r} not found on the header row; available: {sorted(header)}"
        )
    return header[name]


def _is_missing(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip().upper() in _MISSING_TOKENS
    return False


def _optional_rate(value: object, scale: str, *, context: str) -> float | None:
    if _is_missing(value):
        return None
    return apply_rate_scale(scale, to_float(value, context=context), context=context)


def _parse_date(value: object, *, context: str) -> _dt.date:
    """Coerce the workbook's date encodings, `15-Oct-2024` chief among them."""
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if 10_000_000.0 <= number <= 99_991_231.0:              # yyyymmdd
            text = str(int(number))
            return _dt.date(int(text[:4]), int(text[4:6]), int(text[6:8]))
        if 0.0 < number < 200_000.0:                            # Excel serial
            return _dt.date(1899, 12, 30) + _dt.timedelta(days=int(round(number)))
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return _dt.datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    raise ValidationFailure(
        f"{context}: cannot interpret {value!r} as a date (expected 15-Oct-2024, "
        f"mm/dd/yyyy, yyyy-mm-dd, yyyymmdd, or an Excel serial)"
    )


def _optional_date(value: object, *, context: str) -> _dt.date | None:
    return None if _is_missing(value) else _parse_date(value, context=context)


def _resolve_variability_column(header: Mapping[str, int], spec: LoansSheetSpec,
                                census: LoaderCensus, context: str) -> int:
    if spec.col_variability is not None:
        return _column(header, spec.col_variability, context)
    for candidate in _VARIABILITY_ALIASES:
        if candidate in header:
            if candidate != _VARIABILITY_ALIASES[0]:
                census.column_substitutions.append(
                    f"rate-variability column read as {candidate!r} (the sheet's usual spelling is "
                    f"{_VARIABILITY_ALIASES[0]!r})"
                )
            return header[candidate]
    raise ValidationFailure(
        f"{context}: none of the rate-variability column names {list(_VARIABILITY_ALIASES)} were "
        f"found; available: {sorted(header)}. Declare the exact header in config rather than "
        f"letting the loader guess."
    )


def load_facilities(spec: LoansSheetSpec) -> tuple[list[LoanFacility], LoaderCensus]:
    """Read `CORP H.1` into canonical facilities.

    Rows whose Fed Category has no H.1 code cannot appear here by construction —
    Categories 9, 10 and 11 carry no code at all and arrive via
    `load_merged_bucket_balance` instead (PID-LOAN-10)."""
    workbook = _open(spec.workbook)
    try:
        rows = [list(row) for row in _sheet(workbook, spec.h1_sheet, spec.workbook).values]
    finally:
        workbook.close()

    context = f"{spec.workbook.name}:{spec.h1_sheet}"
    census = LoaderCensus()
    header = _header_index(rows, spec.h1_header_row, context)

    idx_id = _column(header, spec.col_facility_id, context)
    # Fallback identifier columns are OPTIONAL: looked up if present, no error if
    # the sheet lacks them (their absence is reported only if it ends up mattering).
    idx_internal = header.get(spec.col_internal_id)
    idx_original = header.get(spec.col_original_internal_id)
    idx_outstanding = header.get(spec.col_outstanding)
    idx_value = header.get(spec.col_value)
    idx_code = _column(header, spec.col_h1_code, context)
    idx_var = _resolve_variability_column(header, spec, census, context)
    idx_locom = _column(header, spec.col_locom, context)
    idx_rate = _column(header, spec.col_interest_rate, context)
    idx_committed = _column(header, spec.col_committed, context)
    idx_utilized = _column(header, spec.col_utilized, context)
    idx_floor = _column(header, spec.col_floor, context)
    idx_orig = _column(header, spec.col_origination, context)
    idx_mat = _column(header, spec.col_maturity, context)

    def cell(row: Sequence[object], index: int) -> object:
        return row[index] if index < len(row) else None

    facilities: list[LoanFacility] = []
    for offset, row in enumerate(rows[spec.h1_header_row:], start=spec.h1_header_row + 1):
        if all(_is_missing(value) for value in row):
            continue
        where = f"{context} row {offset}"
        # Identifier chain (PID-LOAN-12): Customer ID -> Internal ID -> Original
        # Internal ID -> synthesized row label. Rows with [NULL] in all three
        # still carry real balances, so they are LABELED and kept, never dropped —
        # dropping would silently understate segment shares, pool rates, and wt
        # denominators. The synthesized label is unmistakably not a customer ID
        # and names the sheet row, so the record is findable in Excel.
        facility_id: str | None = None
        id_source = "synthesized"
        for candidate_index, source in (
            (idx_id, "customer_id"),
            (idx_internal, "internal_id"),
            (idx_original, "original_internal_id"),
        ):
            if candidate_index is None:
                continue
            candidate = cell(row, candidate_index)
            if not _is_missing(candidate):
                facility_id, id_source = str(candidate).strip(), source
                break
        if facility_id is None:
            facility_id = f"UNIDENTIFIED-ROW-{offset}"
        census.id_sources[id_source] += 1

        raw_code, raw_var, raw_locom = cell(row, idx_code), cell(row, idx_var), cell(row, idx_locom)
        segment = decode_segment(raw_code, raw_var, raw_locom)
        census.reference_keys[reference_key(raw_code, raw_var, raw_locom)] += 1

        committed = apply_money_scale(
            spec.exposure_scale, to_float(cell(row, idx_committed), context=f"{where}: committed"),
            context=f"{where}: committed",
        )
        utilized = apply_money_scale(
            spec.exposure_scale, to_float(cell(row, idx_utilized), context=f"{where}: utilized"),
            context=f"{where}: utilized",
        )
        rate = _optional_rate(cell(row, idx_rate), spec.rate_scale, context=f"{where}: interest rate")
        outstanding_index = idx_value if parse_locom(raw_locom) in (1, 2) else idx_outstanding
        outstanding: float | None = None
        if outstanding_index is not None:
            raw_outstanding = cell(row, outstanding_index)
            if not _is_missing(raw_outstanding):
                outstanding = apply_money_scale(
                    spec.exposure_scale,
                    to_float(raw_outstanding, context=f"{where}: outstanding"),
                    context=f"{where}: outstanding",
                )
        if outstanding is None:
            census.missing_outstanding += 1
            census.missing_outstanding_rows.append((offset, str(facility_id).strip()))
        floor = _optional_rate(cell(row, idx_floor), spec.rate_scale, context=f"{where}: floor")
        originated = _optional_date(cell(row, idx_orig), context=f"{where}: origination date")
        matures = _optional_date(cell(row, idx_mat), context=f"{where}: maturity date")

        census.rows_read += 1
        if id_source == "synthesized":
            census.unidentified_rows.append(offset)
            census.unidentified_exposure += committed
        if rate is None:
            census.missing_interest_rate += 1
            census.missing_interest_rate_exposure += committed
        if floor is None:
            census.missing_floor += 1
        if originated is None:
            census.missing_origination_date += 1
        if matures is None:
            census.missing_maturity_date += 1

        facilities.append(
            LoanFacility(
                facility_id=facility_id,
                segment=segment,
                committed_exposure=committed,
                utilized_exposure=utilized,
                interest_rate=rate,
                interest_rate_floor=floor,
                origination_date=originated,
                maturity_date=matures,
                outstanding_balance=outstanding,
                h1_code=parse_h1_code(raw_code),
            )
        )

    if not facilities:
        raise ValidationFailure(f"{context}: no data rows found below header row {spec.h1_header_row}")
    absent_share_columns = [
        name for name, index in
        ((spec.col_outstanding, idx_outstanding), (spec.col_value, idx_value))
        if index is None
    ]
    if absent_share_columns:
        census.warnings.append(
            f"share-basis column(s) {absent_share_columns} not on header row {spec.h1_header_row} — "
            f"every row on that side will refuse under share_basis='outstanding'. If the sheet "
            f"names them differently, set col_outstanding / col_value in [firm_data.loans]."
        )
    if census.unidentified_rows:
        absent = [
            name for name, index in
            ((spec.col_internal_id, idx_internal), (spec.col_original_internal_id, idx_original))
            if index is None
        ]
        if absent:
            census.warnings.append(
                f"{len(census.unidentified_rows)} row(s) took a synthesized label while fallback "
                f"ID column(s) {absent} are not on the header row — if the sheet carries them "
                f"under different names, configure col_internal_id / col_original_internal_id"
            )
    return facilities, census


def load_merged_bucket_balance(spec: LoansSheetSpec) -> tuple[float, Mapping[str, float]]:
    """Read the Fed Category 9/10/11 balance from the FR Y-9C extract (PID-LOAN-10).

    Those three portfolios carry no H.1 code, so their balance comes from the two
    MDRM line items and they are modelled as ONE merged bucket. Returns the total
    in canonical USD millions plus the per-MDRM parts for the census."""
    workbook = _open(spec.workbook)
    try:
        rows = [list(row) for row in _sheet(workbook, spec.fry9c_sheet, spec.workbook).values]
    finally:
        workbook.close()

    context = f"{spec.workbook.name}:{spec.fry9c_sheet}"
    header = _header_index(rows, spec.fry9c_header_row, context)
    if spec.fry9c_id_column not in header:
        raise ValidationFailure(
            f"{context}: expected {spec.fry9c_id_column!r} on header row {spec.fry9c_header_row}; "
            f"available: {sorted(header)}"
        )
    idx_id = header[spec.fry9c_id_column]
    idx_value = spec.fry9c_value_column_index - 1

    parts: dict[str, float] = {}
    for row in rows[spec.fry9c_header_row:]:
        if idx_id >= len(row):
            continue
        key = row[idx_id]
        if not isinstance(key, str):
            continue
        token = key.strip().upper()
        if token not in MERGED_BUCKET_MDRMS:
            continue
        if idx_value >= len(row):
            raise ValidationFailure(
                f"{context}: {token} has no value in column {spec.fry9c_value_column_index}"
            )
        parts[token] = apply_money_scale(
            spec.fry9c_scale, to_float(row[idx_value], context=f"{context}: {token}"),
            context=f"{context}: {token}",
        )

    missing = [code for code in MERGED_BUCKET_MDRMS if code not in parts]
    if missing:
        raise ValidationFailure(
            f"{context}: MDRM line item(s) {missing} not found under {spec.fry9c_id_column!r}. "
            f"The merged 9/10/11 bucket has no other source, so this is refused rather than "
            f"treated as a zero balance."
        )
    return sum(parts.values()), parts


def load_category_balances(
    spec: LoansSheetSpec,
) -> tuple[Mapping[str, float], Mapping[tuple[str, str], float], LoaderCensus]:
    """Read the per-Fed-Category portfolio balance from `M.1 Balance`.

    The sheet carries its own FRB NII model role per row — column A for the
    domestic side, column B for the international one — so the FR Y-9C line to
    Fed Category wiring is in the workbook and nothing has to be supplied.

    Domestic value columns are attributed to column A's role and international
    ones to column B's, which is how a single FR Y-9C line feeds two categories:
    "c. Secured by farmland" sends its domestic balance to Domestic farmland and
    its international balance to International farmland. Retail and Wholesale-CRE
    rows belong to other models and are skipped.

    Returns (per-category totals, per-(category, side) balances, census). The
    side split matters because the reference results are per category x LOCOM
    block and each block's balance base is the M.1 balance of ITS side: within
    each dom/int column pair, the FIRST value column is HFI-at-AC and the
    SECOND is HFS/FVO (observed layout, 2026-08-07 screenshot).

    Cross-check available to the reader: these totals should reproduce the
    `Sch M bal` column on the FRB SCALARS sheet."""
    workbook = _open(spec.workbook)
    try:
        rows = [list(row) for row in _sheet(workbook, spec.m1_sheet, spec.workbook).values]
    finally:
        workbook.close()

    context = f"{spec.workbook.name}:{spec.m1_sheet}"
    census = LoaderCensus()
    totals: dict[int, float] = {}
    by_side: dict[tuple[int, str], float] = {}
    _SIDE_BY_POSITION = ("HFI", "FVO_HFS")   # first column of each pair = HFI at AC
    skipped = 0

    sides = (
        (spec.m1_domestic_role_column_index - 1, spec.m1_domestic_value_column_indices),
        (spec.m1_international_role_column_index - 1, spec.m1_international_value_column_indices),
    )

    for offset, row in enumerate(rows[spec.m1_first_data_row - 1:], start=spec.m1_first_data_row):
        for role_index, value_indices in sides:
            if role_index >= len(row) or _is_missing(row[role_index]):
                continue
            category = m1_role_category(row[role_index])
            if category is None:
                skipped += 1
                continue
            for position, value_index in enumerate(value_indices):
                cell = row[value_index - 1] if value_index - 1 < len(row) else None
                if _is_missing(cell):
                    continue
                where = f"{context} row {offset} col {value_index}"
                amount = apply_money_scale(
                    spec.m1_scale, to_float(cell, context=where), context=where
                )
                totals[category] = totals.get(category, 0.0) + amount
                side = _SIDE_BY_POSITION[min(position, 1)]
                by_side[(category, side)] = by_side.get((category, side), 0.0) + amount

    if not totals:
        raise ValidationFailure(
            f"{context}: no 'Wholesale - Corp' roles found from row {spec.m1_first_data_row}. "
            f"Check m1_first_data_row and the role column indices before trusting a zero balance."
        )

    missing = sorted(set(FED_CATEGORY_NAMES) - set(totals))
    if missing:
        census.warnings.append(
            f"M.1 supplied no balance for Fed Category {missing} "
            f"({[FED_CATEGORY_NAMES[c] for c in missing]}) — those categories project zero income"
        )
    census.rows_read = len(totals)
    census.warnings.append(f"skipped {skipped} non-Corporate role cells (Retail and Wholesale-CRE)")
    return (
        MappingProxyType({FED_CATEGORY_NAMES[c]: v for c, v in totals.items()}),
        MappingProxyType({(FED_CATEGORY_NAMES[c], side): v for (c, side), v in by_side.items()}),
        census,
    )


_H2_CODE_ALIASES = ("Line Reported on FR Y-9C", "Line Reported on FR Y9C")


def load_cre_facilities(spec: LoansSheetSpec) -> tuple[list["LoanFacility"], LoaderCensus]:
    """Read the CRE H.2 sheet into canonical facilities (PID-LOAN-18/19).

    Differences from the Corporate H.1 read, each PID-backed:

    - the balance columns are `Committed Balance` and `Outstanding Balance`
      (H.2's own names); `Utilized Exposure Global` does not exist on H.2, so
      the facility's utilized field is carried as 0 and nothing may weight by it
    - a blank Outstanding Balance is a genuine ZERO (an undrawn facility), not
      a refusal — counted in the census
    - rows whose line code is DO NOT USE (codes 4 and 6) belong to no CRE
      category: excluded and censused with their exposure, never decoded
    - identifier columns are OPTIONAL on H.2 (the PID-LOAN-12 chain applies
      where present; otherwise rows are labeled by sheet row and censused)"""
    from ..interest_income.loans_schemas import LoanFacility as _LoanFacility
    from .loans_cre_mapping import (
        H2_DO_NOT_USE_CODES,
        cre_reference_key,
        decode_cre_segment,
        parse_h2_code,
    )

    if spec.cre_h2_sheet is None:
        raise ValidationFailure(
            "cre_h2_sheet is not configured — the CRE run is enabled by naming the H.2 sheet "
            "in [firm_data.loans]"
        )
    workbook = _open(spec.workbook)
    try:
        rows = [list(row) for row in _sheet(workbook, spec.cre_h2_sheet, spec.workbook).values]
    finally:
        workbook.close()

    context = f"{spec.workbook.name}:{spec.cre_h2_sheet}"
    census = LoaderCensus(title="CRE LOANS LOADER CENSUS")
    header = _header_index(rows, spec.cre_h2_header_row, context)

    idx_code = None
    for candidate in _H2_CODE_ALIASES:
        if candidate in header:
            if candidate != _H2_CODE_ALIASES[0]:
                census.column_substitutions.append(
                    f"line-code column read as {candidate!r} (the H.2 header usually spells it "
                    f"{_H2_CODE_ALIASES[0]!r})"
                )
            idx_code = header[candidate]
            break
    if idx_code is None:
        raise ValidationFailure(
            f"{context}: none of the line-code column names {list(_H2_CODE_ALIASES)} were found; "
            f"available: {sorted(header)}"
        )
    idx_var = _resolve_variability_column(header, spec, census, context)
    idx_locom = _column(header, spec.col_locom, context)
    idx_rate = _column(header, spec.col_interest_rate, context)
    idx_committed = _column(header, spec.cre_col_committed, context)
    idx_outstanding = _column(header, spec.cre_col_outstanding, context)
    idx_value = _column(header, spec.cre_col_value, context) if spec.cre_col_value else None
    idx_floor = _column(header, spec.col_floor, context)
    idx_orig = _column(header, spec.col_origination, context)
    idx_mat = _column(header, spec.col_maturity, context)
    # identifier chain, all optional on H.2
    idx_ids = [
        (header.get(spec.col_facility_id), "customer_id"),
        (header.get(spec.col_internal_id), "internal_id"),
        (header.get(spec.col_original_internal_id), "original_internal_id"),
    ]

    def cell(row: Sequence[object], index: int) -> object:
        return row[index] if index < len(row) else None

    facilities: list = []
    for offset, row in enumerate(rows[spec.cre_h2_header_row:], start=spec.cre_h2_header_row + 1):
        if all(_is_missing(value) for value in row):
            continue
        where = f"{context} row {offset}"

        raw_code = cell(row, idx_code)
        code = parse_h2_code(raw_code)
        committed = apply_money_scale(
            spec.exposure_scale, to_float(cell(row, idx_committed), context=f"{where}: committed"),
            context=f"{where}: committed",
        )
        if code in H2_DO_NOT_USE_CODES:
            census.excluded_line_codes[code] += 1
            census.excluded_line_code_exposure += committed
            continue

        facility_id: str | None = None
        id_source = "synthesized"
        for candidate_index, source in idx_ids:
            if candidate_index is None:
                continue
            candidate = cell(row, candidate_index)
            if not _is_missing(candidate):
                facility_id, id_source = str(candidate).strip(), source
                break
        if facility_id is None:
            facility_id = f"UNIDENTIFIED-ROW-{offset}"
        census.id_sources[id_source] += 1

        raw_var, raw_locom = cell(row, idx_var), cell(row, idx_locom)
        segment = decode_cre_segment(raw_code, raw_var, raw_locom)
        census.reference_keys[cre_reference_key(raw_code, raw_var, raw_locom)] += 1

        # HFS/FVO rows weight by the distinct Value column where one is
        # configured (the "Launchpoint Value" residual); otherwise Outstanding
        # Balance weights both sides.
        outstanding_index = idx_outstanding
        if idx_value is not None and parse_locom(raw_locom) in (1, 2):
            outstanding_index = idx_value
        raw_outstanding = cell(row, outstanding_index)
        if _is_missing(raw_outstanding):
            outstanding = 0.0
            census.blank_outstanding += 1
        else:
            outstanding = apply_money_scale(
                spec.exposure_scale,
                to_float(raw_outstanding, context=f"{where}: outstanding"),
                context=f"{where}: outstanding",
            )

        rate = _optional_rate(cell(row, idx_rate), spec.rate_scale, context=f"{where}: interest rate")
        floor = _optional_rate(cell(row, idx_floor), spec.rate_scale, context=f"{where}: floor")
        originated = _optional_date(cell(row, idx_orig), context=f"{where}: origination date")
        matures = _optional_date(cell(row, idx_mat), context=f"{where}: maturity date")

        census.rows_read += 1
        if id_source == "synthesized":
            census.unidentified_rows.append(offset)
            census.unidentified_exposure += committed
        if rate is None:
            census.missing_interest_rate += 1
            census.missing_interest_rate_exposure += committed
        if floor is None:
            census.missing_floor += 1
        if originated is None:
            census.missing_origination_date += 1
        if matures is None:
            census.missing_maturity_date += 1

        facilities.append(
            _LoanFacility(
                facility_id=facility_id,
                segment=segment,
                committed_exposure=committed,
                utilized_exposure=0.0,        # H.2 carries no utilized column; nothing weights by it
                interest_rate=rate,
                interest_rate_floor=floor,
                origination_date=originated,
                maturity_date=matures,
                outstanding_balance=outstanding,
                h1_code=None,                 # H.2 codes are a different vocabulary; not carried here
            )
        )

    if not facilities:
        raise ValidationFailure(
            f"{context}: no CRE data rows found below header row {spec.cre_h2_header_row} "
            f"(DO-NOT-USE rows excluded: {sum(census.excluded_line_codes.values())})"
        )
    return facilities, census


def load_cre_side_balances(
    spec: LoansSheetSpec,
) -> tuple[Mapping[tuple[str, str], float], list[str]]:
    """Read the CRE Eq A32 multiplicand from the configured M.1 rows (PID-LOAN-20).

    Domestic categories take the E/G (HFI / HFS-FVO) values of their own row —
    construction, multifamily, non-owner-occupied — and the merged international
    category sums the I/K columns over the SAME three rows (the three blue
    "Wholesale - CRE - international" role labels of the user's screenshot).
    Blank cells are genuine zeros here: a side with no balance is real (the
    workbook's international HFS/FVO side is empty), not a refusal.

    The wiring is by row number because the user confirmed the cells directly;
    the role labels on those rows are returned as notes so a misconfigured row
    is visible — a configured row whose domestic role does not look like a
    Wholesale-CRE label earns a warning, never a silent sum."""
    from .loans_cre_mapping import CRE_CATEGORY_NAMES, M1_CRE_ROLE_PREFIX
    from .loans_mapping import normalize_role

    workbook = _open(spec.workbook)
    try:
        rows = [list(row) for row in _sheet(workbook, spec.m1_sheet, spec.workbook).values]
    finally:
        workbook.close()
    context = f"{spec.workbook.name}:{spec.m1_sheet}"

    configured = (
        (1, spec.cre_m1_construction_row),
        (2, spec.cre_m1_multifamily_row),
        (3, spec.cre_m1_non_owner_occupied_row),
    )
    dom_hfi_col, dom_hfs_col = spec.m1_domestic_value_column_indices
    intl_hfi_col, intl_hfs_col = spec.m1_international_value_column_indices

    def amount(row: Sequence[object], column_index: int, where: str) -> float:
        cell = row[column_index - 1] if column_index - 1 < len(row) else None
        if _is_missing(cell):
            return 0.0
        return apply_money_scale(spec.m1_scale, to_float(cell, context=where), context=where)

    balances: dict[tuple[str, str], float] = {}
    notes: list[str] = []
    for category, row_number in configured:
        if row_number > len(rows):
            raise ValidationFailure(
                f"{context}: configured CRE M.1 row {row_number} is beyond the end of the sheet"
            )
        row = rows[row_number - 1]
        name = CRE_CATEGORY_NAMES[category]
        where = f"{context} row {row_number}"
        balances[(name, "HFI")] = balances.get((name, "HFI"), 0.0) + amount(row, dom_hfi_col, where)
        balances[(name, "FVO_HFS")] = (
            balances.get((name, "FVO_HFS"), 0.0) + amount(row, dom_hfs_col, where)
        )
        intl_name = CRE_CATEGORY_NAMES[4]
        balances[(intl_name, "HFI")] = (
            balances.get((intl_name, "HFI"), 0.0) + amount(row, intl_hfi_col, where)
        )
        balances[(intl_name, "FVO_HFS")] = (
            balances.get((intl_name, "FVO_HFS"), 0.0) + amount(row, intl_hfs_col, where)
        )

        role_dom = row[spec.m1_domestic_role_column_index - 1] if spec.m1_domestic_role_column_index - 1 < len(row) else None
        role_intl = row[spec.m1_international_role_column_index - 1] if spec.m1_international_role_column_index - 1 < len(row) else None
        notes.append(f"row {row_number} ({name}): roles dom={role_dom!r} intl={role_intl!r}")
        for side, role in (("domestic", role_dom), ("international", role_intl)):
            if role is not None and not normalize_role(role).startswith(M1_CRE_ROLE_PREFIX):
                notes.append(
                    f"WARN: M.1 row {row_number} {side} role {role!r} does not look like a "
                    f"'Wholesale - CRE' label — check the cre_m1_*_row configuration "
                    f"(PID-LOAN-20 wires rows 17/18/21 on the company sheet)"
                )
    return MappingProxyType(balances), notes


def load_3m_treasury(
    spec: LoansSheetSpec, scenario: str, quarters: Sequence[int], launch_point: str
) -> tuple[Mapping[str, float], Mapping[int, float], float]:
    """Read the 3-month Treasury history and projection from `MEV`.

    History rows carry the scenario name `Actual` and are keyed by calendar
    quarter for the median-origination lookup; projection rows carry `scenario`
    and are mapped onto PQ1..PQn in sheet order. Also returns the launch-point
    value, read from the history at `launch_point` (e.g. `2024Q4`)."""
    workbook = _open(spec.workbook)
    try:
        rows = [list(row) for row in _sheet(workbook, spec.mev_sheet, spec.workbook).values]
    finally:
        workbook.close()

    context = f"{spec.workbook.name}:{spec.mev_sheet}"
    header = _header_index(rows, spec.mev_header_row, context)
    idx_scenario = _column(header, spec.mev_scenario_column, context)
    idx_date = _column(header, spec.mev_date_column, context)
    idx_rate = _column(header, spec.mev_3m_column, context)

    launch_label = check_quarter_label(f"{context}: launch point", launch_point)
    launch_year, launch_quarter = int(launch_label[:4]), int(launch_label[5])

    def quarters_since_launch(label: str) -> int:
        return (int(label[:4]) - launch_year) * 4 + int(label[5]) - launch_quarter

    history: dict[str, float] = {}
    projection: dict[int, float] = {}
    seen_scenario_labels: set[str] = set()
    for row in rows[spec.mev_header_row:]:
        if max(idx_scenario, idx_date, idx_rate) >= len(row):
            continue
        name, raw_date, raw_rate = row[idx_scenario], row[idx_date], row[idx_rate]
        if _is_missing(name) or _is_missing(raw_date) or _is_missing(raw_rate):
            continue
        rate = apply_rate_scale(
            spec.mev_rate_scale, to_float(raw_rate, context=f"{context}: 3M"), context=f"{context}: 3M"
        )
        label = check_quarter_label(
            f"{context}: date", str(raw_date).strip().replace(" ", "").upper()
        )
        if str(name).strip() == spec.mev_history_scenario:
            # A duplicated history quarter with a DIFFERENT value is the silent-bug
            # case: dict assignment would keep whichever came last with no trace.
            if label in history and history[label] != rate:
                raise ValidationFailure(
                    f"{context}: {spec.mev_history_scenario!r} carries {label} twice with "
                    f"different values ({history[label]!r} vs {rate!r})"
                )
            history[label] = rate
        elif str(name).strip() == scenario:
            # Projection rows are mapped to PQ indices BY THEIR DATE, never by
            # sheet order: the sheet says which quarter each row is, so use it.
            # Rows outside PQ1..PQn (a 13-quarter supervisory path has a tail we
            # do not need) are simply out of horizon, not errors.
            if label in seen_scenario_labels:
                raise ValidationFailure(f"{context}: scenario {scenario!r} carries {label} twice")
            seen_scenario_labels.add(label)
            index = quarters_since_launch(label)
            if index in quarters:
                projection[index] = rate

    if not history:
        raise ValidationFailure(
            f"{context}: no rows under scenario {spec.mev_history_scenario!r} — the fixed-rate "
            f"spread needs the 3M history back to the earliest median origination quarter"
        )
    missing = [q for q in quarters if q not in projection]
    if missing:
        expected = [
            f"PQ{q} ({launch_year + (launch_quarter + q - 1) // 4}"
            f"Q{(launch_quarter + q - 1) % 4 + 1})"
            for q in missing
        ]
        available = sorted(seen_scenario_labels) or ["<none>"]
        raise ValidationFailure(
            f"{context}: scenario {scenario!r} is missing {', '.join(expected)} relative to "
            f"launch point {launch_label}; scenario rows found: {', '.join(available)}. "
            f"Check the --scenario spelling against the sheet and the launch point."
        )
    if launch_point not in history:
        raise ValidationFailure(
            f"{context}: launch point {launch_point!r} is absent from the {spec.mev_history_scenario!r} "
            f"history — the floating spread is measured against it"
        )

    return history, projection, history[launch_point]


# --- Reference results (compare mode) --------------------------------------

_BLOCK_MARKER = re.compile(r"^\s*(\d{1,2})\s*-\s*(HFI|HFS/FVO)\s*$")
_MERGED_MARKER = re.compile(r"^\s*9\s*,\s*10\s*,\s*11\s*$")
_RESULT_ROW_LABELS = {
    "Fixed Income": "fixed",
    "Variable Rate Income": "variable",
    "Total": "total",
}


def load_reference_results(
    spec: LoansSheetSpec,
    sheet: str | None = None,
    category_names: Mapping[int, str] | None = None,
    include_merged: bool = True,
) -> Mapping[tuple[str, str], Mapping[str, Mapping[int, float]]]:
    """Read the workbook's own projected income for the compare mode.

    Layout (observed 2026-08-12, identical for Corporate and CRE): blocks
    introduced by a cell reading `N - HFI` or `N - HFS/FVO` (Corporate adds one
    `9, 10, 11` merged block), each carrying rows labelled `Fixed Income`,
    `Variable Rate Income`, and `Total` over PQ0..PQ9 columns headed
    `PQ0`..`PQ9`, values in raw dollars ("-" meaning zero).

    `sheet` defaults to the Corporate results sheet; `category_names` maps the
    block marker's number to a category name (default: the eleven Corporate Fed
    Categories — pass CRE_CATEGORY_NAMES for the CRE sheet, whose markers are
    the four workbook categories); `include_merged` recognizes the Corporate
    `9, 10, 11` marker and is off for CRE, which has no merged block.

    Returns {(category name, class): {"fixed"|"variable"|"total": {0..9: USD
    millions}}}. A Corporate block's Total exceeds Fixed + Variable — the
    workbook sums a third stream (Mixed, by its rate signature); the compare
    derives it as Total − Fixed − Variable rather than requiring its row."""
    from .loans_mapping import CLASS_FVO_HFS, CLASS_HFI, FED_CATEGORY_NAMES

    sheet_name = sheet if sheet is not None else spec.results_sheet
    names = category_names if category_names is not None else FED_CATEGORY_NAMES
    if sheet_name is None:
        raise ValidationFailure("results_sheet is not configured — nothing to compare against")

    workbook = _open(spec.workbook)
    try:
        rows = [list(row) for row in _sheet(workbook, sheet_name, spec.workbook).values]
    finally:
        workbook.close()
    context = f"{spec.workbook.name}:{sheet_name}"

    quarter_columns: dict[int, int] = {}
    for row in rows[:20]:
        found = {
            int(str(cell).strip()[2:]): index
            for index, cell in enumerate(row)
            if isinstance(cell, str) and re.fullmatch(r"PQ[0-9]", cell.strip())
        }
        if len(found) >= 10:
            quarter_columns = found
            break
    if not quarter_columns:
        raise ValidationFailure(
            f"{context}: no header row carrying PQ0..PQ9 labels found in the first 20 rows"
        )

    class_by_marker = {"HFI": CLASS_HFI, "HFS/FVO": CLASS_FVO_HFS}
    results: dict[tuple[str, str], dict[str, Mapping[int, float]]] = {}
    current: tuple[str, str] | None = None

    for row in rows:
        for cell in row:
            if not isinstance(cell, str):
                continue
            marker = _BLOCK_MARKER.match(cell)
            if marker:
                category = int(marker.group(1))
                if category not in names:
                    raise ValidationFailure(
                        f"{context}: block marker {cell!r} names category {category}, which is "
                        f"not one of {sorted(names)}"
                    )
                current = (names[category], class_by_marker[marker.group(2)])
                break
            if include_merged and _MERGED_MARKER.match(cell):
                current = (FED_CATEGORY_NAMES[9], "MERGED")
                break
        label = next(
            (
                _RESULT_ROW_LABELS[str(cell).strip()]
                for cell in row
                if isinstance(cell, str) and str(cell).strip() in _RESULT_ROW_LABELS
            ),
            None,
        )
        if label is None or current is None:
            continue
        path: dict[int, float] = {}
        for quarter, index in quarter_columns.items():
            cell = row[index] if index < len(row) else None
            if _is_missing(cell) or (isinstance(cell, str) and cell.strip() == "-"):
                path[quarter] = 0.0
            else:
                path[quarter] = to_float(cell, context=f"{context}: {current} {label} PQ{quarter}") / 1e6
        results.setdefault(current, {})[label] = MappingProxyType(path)

    if not results:
        raise ValidationFailure(
            f"{context}: no result blocks found — expected markers like '1 - HFI' with "
            f"Fixed Income / Variable Rate Income / Total rows"
        )
    return MappingProxyType({key: MappingProxyType(streams) for key, streams in results.items()})
