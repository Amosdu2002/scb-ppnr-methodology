"""Physical table readers: CSV (stdlib) and XLSX (optional openpyxl dependency).

A "table" is a header row plus records; values come back untyped (str, number, or
date cell) and are parsed downstream. No workbook path, sheet name, or cell
reference is ever hard-coded here — every physical detail arrives from the
company-local configuration file."""

from __future__ import annotations

import csv
from pathlib import Path

from ..interest_expense.schemas import ValidationFailure

Row = dict[str, object]


def read_table(path: Path, sheet: str | None = None) -> list[Row]:
    if not path.exists():
        raise ValidationFailure(f"input file not found: {path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        if sheet:
            raise ValidationFailure(f"{path}: 'sheet' applies only to workbook formats, not CSV")
        return _read_csv(path)
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx(path, sheet)
    raise ValidationFailure(f"{path}: unsupported format {suffix!r} (use .csv or .xlsx)")


def _read_csv(path: Path) -> list[Row]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValidationFailure(f"{path}: file has no header row")
        rows: list[Row] = []
        for record in reader:
            cleaned: Row = {}
            for key, value in record.items():
                if key is None:
                    continue
                cleaned[key.strip()] = value
            if any(str(v).strip() for v in cleaned.values() if v is not None):
                rows.append(cleaned)
        return rows


def _read_xlsx(path: Path, sheet: str | None) -> list[Row]:
    try:
        import openpyxl
    except ImportError:
        raise ValidationFailure(
            f"{path}: reading .xlsx requires the optional 'openpyxl' dependency "
            f"(pip install openpyxl, or install this package with the [excel] extra)"
        ) from None
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is not None:
            if sheet not in workbook.sheetnames:
                raise ValidationFailure(f"{path}: sheet {sheet!r} not found; available: {workbook.sheetnames}")
            worksheet = workbook[sheet]
        else:
            worksheet = workbook.worksheets[0]
        cells = worksheet.iter_rows(values_only=True)
        header = next(cells, None)
        if header is None:
            raise ValidationFailure(f"{path}: sheet {worksheet.title!r} has no header row")
        names = [str(cell).strip() if cell is not None else "" for cell in header]
        rows: list[Row] = []
        for record in cells:
            if all(cell is None or str(cell).strip() == "" for cell in record):
                continue
            rows.append({name: value for name, value in zip(names, record) if name})
        return rows
    finally:
        workbook.close()
