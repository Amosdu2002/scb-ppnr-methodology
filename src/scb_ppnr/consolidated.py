"""Consolidated results output — one workbook (plus a flat CSV twin) gathering
every family's quarterly paths and the run's headline reconciliation.

Sheet plan (all amounts USD MILLIONS per quarter, pre-hedge — D-006):

  Summary — run identity, the headline income/expense/NII paths against the
            FRB targets, the reconciliation verdicts and the two alpha_b
            calibrations, and each component's nine-quarter cumulative.
  Income  — every component income path: the loans parts when supplied, the
            six siblings, their sum, modeled and implied trading NII, the
            family total, the FRB target, and the per-quarter difference.
  Expense — the five expense model paths, their total, the FRB target, and
            the per-quarter difference. Omitted when the expense family did
            not run (--skip-expense or inputs not configured).

`consolidated_tables` is pure (no I/O, no openpyxl) so content is testable
anywhere. `write_consolidated_workbook` needs openpyxl (lazy import — the
caller decides how to handle its absence). `write_consolidated_csv` is
stdlib-only and carries exactly the quarterly-path rows (label + PQ1..PQ9 +
nine-quarter total); scalar lines — verdicts, alpha values — live in the
workbook Summary and the text reports.
"""

from __future__ import annotations

from pathlib import Path
from typing import Mapping

from .core.schemas import PROJECTION_QUARTERS
from .interest_expense.orchestrator import MODEL_EXECUTION_ORDER
from .interest_income.nii_trading_al import SIBLING_MODEL_IDS

_LOANS_PART_ORDER = ("ii_loans_corporate", "ii_loans_cre", "ii_loans_retail")
PATH_ROW_WIDTH = 1 + len(PROJECTION_QUARTERS) + 1  # label + PQ1..PQ9 + 9Q total


def _path_row(label: str, path: Mapping[int, float]) -> list[object]:
    values = [float(path[q]) for q in PROJECTION_QUARTERS]
    return [label, *values, sum(values)]


def _header_row(first: str) -> list[object]:
    return [first, *(f"PQ{q}" for q in PROJECTION_QUARTERS), "9Q total"]


def consolidated_tables(
    *,
    income_result,
    implied_path: Mapping[int, float],
    loans_parts: Mapping[str, Mapping[int, float]],
    frb_total_interest_income: Mapping[int, float],
    expense_result=None,
    frb_total_interest_expense: Mapping[int, float] | None = None,
    monitor=None,
    generated: str,
) -> dict[str, list[list[object]]]:
    income_total = income_result.total_income_path()

    summary: list[list[object]] = [
        ["scb_ppnr consolidated results"],
        ["firm", income_result.firm_id],
        ["scenario", income_result.scenario_id],
        ["generated", generated],
        ["units", "USD millions per quarter — pre-hedge (D-006)"],
        ["methodology", "Federal Reserve proposed 2026 suite — NOT adopted"],
        [],
        _header_row("HEADLINE"),
        _path_row("total interest income (modeled)", income_total),
        _path_row("frb_total_interest_income (target)", frb_total_interest_income),
        _path_row("income difference (modeled − FRB)",
                  income_result.reconciliation.per_quarter_difference),
    ]
    if expense_result is not None and frb_total_interest_expense is not None:
        expense_total = {
            q: sum(r.expense_path()[q] for r in expense_result.results.values())
            for q in PROJECTION_QUARTERS
        }
        summary += [
            _path_row("total interest expense (modeled)", expense_total),
            _path_row("frb_total_interest_expense (target)", frb_total_interest_expense),
            _path_row("expense difference (modeled − FRB)",
                      expense_result.reconciliation.per_quarter_difference),
        ]
    if monitor is not None:
        summary.append(_path_row("net interest income (modeled)", monitor.nii_path))
        if monitor.frb_net_interest_income is not None:
            summary += [
                _path_row("frb_net_interest_income (target)", monitor.frb_net_interest_income),
                _path_row("nii difference (modeled − FRB; quarterly gaps are structural)",
                          monitor.per_quarter_gap),
            ]

    summary += [
        [],
        ["RECONCILIATION AND CALIBRATION"],
        ["income family cumulative difference (exact by construction, PID-TRD-1)",
         income_result.reconciliation.cumulative_difference],
        ["alpha_b — nii_trading_al (annualized, PID-TRD-3 basis)",
         income_result.calibration.alpha_b],
    ]
    if expense_result is not None:
        summary += [
            ["expense family cumulative difference (exact by construction, PID-OB-5)",
             expense_result.reconciliation.cumulative_difference],
            ["alpha_b — ie_other_borrowing (annualized)",
             expense_result.calibration.alpha_b],
        ]
    if monitor is not None:
        if monitor.within_identity_guard is None:
            summary.append(["combined-NII monitor",
                            "n/a — frb_net_interest_income not supplied"])
        else:
            summary += [
                ["combined-NII monitor cumulative gap", monitor.cumulative_gap],
                ["combined-NII monitor: within 1% identity guard",
                 str(monitor.within_identity_guard)],
            ]

    summary += [[], ["NINE-QUARTER CUMULATIVE BY COMPONENT"], ["component", "9Q total"]]
    for model_id, cumulative in income_result.reconciliation.components_cumulative.items():
        summary.append([model_id, float(cumulative)])
    if expense_result is not None:
        for model_id, cumulative in expense_result.reconciliation.components_cumulative.items():
            summary.append([model_id, float(cumulative)])

    income: list[list[object]] = [
        ["INCOME COMPONENT PATHS (USD millions per quarter)"],
        _header_row("component"),
    ]
    for part in _LOANS_PART_ORDER:
        if part in loans_parts:
            income.append(_path_row(part, loans_parts[part]))
    for model_id in SIBLING_MODEL_IDS:
        income.append(_path_row(model_id, income_result.sibling_paths[model_id]))
    income += [
        _path_row("sum of siblings", {
            q: sum(path[q] for path in income_result.sibling_paths.values())
            for q in PROJECTION_QUARTERS
        }),
        _path_row("nii_trading_al (modeled)", income_result.trading_result.income_path()),
        _path_row("implied trading (round-0 diagnostic)", implied_path),
        _path_row("total interest income", income_total),
        _path_row("frb_total_interest_income (target)", frb_total_interest_income),
        _path_row("difference (seven components − FRB)",
                  income_result.reconciliation.per_quarter_difference),
    ]

    tables = {"Summary": summary, "Income": income}

    if expense_result is not None and frb_total_interest_expense is not None:
        expense: list[list[object]] = [
            ["EXPENSE MODEL PATHS (USD millions per quarter)"],
            _header_row("model"),
        ]
        for model_id in MODEL_EXECUTION_ORDER:
            expense.append(_path_row(model_id, expense_result.results[model_id].expense_path()))
        expense += [
            _path_row("total interest expense", {
                q: sum(r.expense_path()[q] for r in expense_result.results.values())
                for q in PROJECTION_QUARTERS
            }),
            _path_row("frb_total_interest_expense (target)", frb_total_interest_expense),
            _path_row("difference (five components − FRB)",
                      expense_result.reconciliation.per_quarter_difference),
        ]
        tables["Expense"] = expense

    return tables


def _is_path_row(row: list[object]) -> bool:
    return (len(row) == PATH_ROW_WIDTH and isinstance(row[0], str)
            and all(isinstance(cell, (int, float)) and not isinstance(cell, bool)
                    for cell in row[1:]))


def write_consolidated_csv(tables: Mapping[str, list[list[object]]], path: Path) -> None:
    """Flat, diff-friendly twin: every quarterly-path row from every sheet,
    prefixed by its sheet name. %.6f matches the --paths-out convention."""
    lines = ["sheet,series," + ",".join(f"PQ{q}" for q in PROJECTION_QUARTERS) + ",total_9q"]
    for sheet, rows in tables.items():
        for row in rows:
            if _is_path_row(row):
                label = str(row[0]).replace(",", ";")
                lines.append(f"{sheet},{label}," + ",".join(f"{cell:.6f}" for cell in row[1:]))
    Path(path).write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_consolidated_workbook(tables: Mapping[str, list[list[object]]], path: Path) -> None:
    """Requires openpyxl (lazy import; caller handles ImportError)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    book = Workbook()
    book.remove(book.active)
    bold = Font(bold=True)
    for sheet_name, rows in tables.items():
        sheet = book.create_sheet(sheet_name)
        for row in rows:
            sheet.append(row if row else [None])
        for excel_row in sheet.iter_rows():
            if not any(isinstance(cell.value, float) for cell in excel_row):
                for cell in excel_row:
                    if cell.value is not None:
                        cell.font = bold
            for cell in excel_row:
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.000"
        sheet.column_dimensions["A"].width = 52
        for index in range(2, PATH_ROW_WIDTH + 1):
            sheet.column_dimensions[get_column_letter(index)].width = 12
        sheet.freeze_panes = "B2"
    book.save(path)
