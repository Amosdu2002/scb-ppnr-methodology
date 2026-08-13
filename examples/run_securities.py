"""Securities family (ii_ust / ii_mbs / ii_other_sec) driven by a config file.

    PYTHONPATH=src python3 examples/run_securities.py                       # synthetic demo
    PYTHONPATH=src python3 examples/run_securities.py \
        --config config/local/company.toml --scenario severely_adverse      # company run

The company config needs a [firm_data.securities] section (PID-SEC-6 — see
config/company.template.toml). The synthetic demo generates a small in-memory
workbook (invented securities, hand-calculable numbers) so no .xlsx is committed.
Requires openpyxl. Output amounts: USD MILLIONS per quarter (D-006), pre-hedge."""

from __future__ import annotations

import argparse
import datetime as dt
import tempfile
from pathlib import Path

from scb_ppnr.ingestion import load_config, load_mev_scenario, load_securities_inputs
from scb_ppnr.ingestion.config import (
    FirmDataConfig,
    IngestionConfig,
    SecuritiesConfig,
    SecuritiesEnrichmentSheet,
    TableSource,
)
from scb_ppnr.interest_income import (
    FLOOR_MODE_SECURITY,
    IncomeModelResult,
    IncomeScenarioPaths,
    project_mbs,
    project_other_sec,
    project_ust,
)
from scb_ppnr.interest_income.schemas import PROJECTION_QUARTERS


_WARNING_BUCKETS = (
    "floored at", "negative imputed margin", "straight-line form", "step-coupon",
    "PID-SEC-3", "PID-SEC-7", "multi-family", "no enrichment match",
    "skipped (on_security_error", "apportioned", "read as 0", "PQ1 face is 0",
    "floater-indicator", "skipped on error",
)


def _render_model(result: IncomeModelResult) -> str:
    lines = [f"\n{result.model_id}  [{result.validation_status}]"]
    lines.append("  quarter        " + "  ".join(f"   PQ{q}" for q in PROJECTION_QUARTERS))
    lines.append("  income         " + "  ".join(f"{result.income_path()[q]:7.2f}" for q in PROJECTION_QUARTERS))
    diag = [result.quarters[q - 1].diagnostics for q in PROJECTION_QUARTERS]
    lines.append("  coupon leg     " + "  ".join(f"{d.coupon_accrual:7.2f}" for d in diag))
    lines.append("  accretion leg  " + "  ".join(f"{d.accretion:7.2f}" for d in diag))
    lines.append("  reinvested     " + "  ".join(f"{d.reinvested_income:7.2f}" for d in diag))
    lines.append(f"  9-quarter cumulative income: {result.cumulative_income:.2f}")
    return "\n".join(lines)


def _warning_summary(all_warnings: list[str], show: int) -> str:
    from collections import Counter, defaultdict
    buckets: Counter[str] = Counter()
    examples: dict[str, list[str]] = defaultdict(list)
    for warning in all_warnings:
        key = next((b for b in _WARNING_BUCKETS if b in warning), "other")
        buckets[key] += 1
        if len(examples[key]) < show:
            examples[key].append(warning)
    lines = [f"\nWARNING SUMMARY ({len(all_warnings)} total — use --warnings full or --report FILE for every line):"]
    for key, count in buckets.most_common():
        lines.append(f"  [{count:>6}] {key}")
        for example in examples[key][: 2 if count > show else show]:
            lines.append(f"           e.g. {example}")
    highlights = [w for w in all_warnings if w.startswith("HIGHLIGHT")]
    if highlights:
        lines.append(f"\nHIGHLIGHTS ({len(highlights)}):")
        lines.extend(f"  {w}" for w in highlights[:30])
        if len(highlights) > 30:
            lines.append(f"  … {len(highlights) - 30} more (see --report FILE)")
    return "\n".join(lines)


def _synthetic_demo() -> tuple[IngestionConfig, IncomeScenarioPaths]:
    import openpyxl

    tmp = Path(tempfile.mkdtemp(prefix="scb_ppnr_securities_demo_"))
    epoch = dt.date(1899, 12, 30)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Positions"
    mdrm = ["D_DT", "CQSCP083", "CQSCP084", "CQSCP087", "CQSCP089", "CQSCP090", "CQSCP092", "CQSCP094", "CQSCJH21"]
    for filler in (["_"] * len(mdrm), ["tech"] * len(mdrm), ["desc"] * len(mdrm)):
        ws.append(filler)
    ws.append(mdrm)
    ws.append([20241231, "DEMOUST01", "US Treasuries & Agencies", 960e6, 1000e6, 1000e6, "AFS", 4.2, None])
    ws.append([20241231, "DEMOAGY01", "Agency MBS", 950e6, 1000e6, 1200e6, "AFS", 5.0, None])
    ws.append([20241231, "DEMOOTH01", "Corporate Bond", 2000e6, 2000e6, 2000e6, "AFS", 5.0, None])
    ito = wb.create_sheet("ITO")
    ito.append(["CUSIP", "Maturity Date", "Coupon Rate", "CPN_TYP", "CPN_Floor", "WAL"])
    ito.append(["DEMOUST01", (dt.date(2025, 12, 31) - epoch).days, 4.0, "FIXED", None, None])
    ito.append(["DEMOAGY01", None, 5.0, "FIXED", None, 2.5])
    ito.append(["DEMOOTH01", (dt.date(2035, 6, 30) - epoch).days, 5.0, "FIXED", None, None])
    prepay = wb.create_sheet("prepay")
    prepay.append(["Sum of current_face", "Column Labels"])
    prepay.append(["Row Labels", 0, 3, 6, 9, 12, 15, 18, 21, 24, 27, "Grand Total"])
    prepay.append(["DEMOAGY01", 1000e6, 800e6, 800e6, 800e6, 800e6, 800e6, 800e6, 800e6, 800e6, 800e6, 0])
    wb.save(tmp / "securities.xlsx")

    config = IngestionConfig(
        base_dir=tmp,
        firm_data=FirmDataConfig(
            firm_id="SYNTHETIC_FIRM",
            spot=TableSource(Path("unused.csv")),
            quarterly=TableSource(Path("unused.csv")),
            securities=SecuritiesConfig(
                workbook=Path("securities.xlsx"), positions_sheet="Positions",
                money_scale="dollars", coupon_scale="percent", book_yield_scale="percent",
                floor_mode=FLOOR_MODE_SECURITY, prepayment_sheet="prepay",
                enrichment=(SecuritiesEnrichmentSheet(
                    sheet="ITO", key_column="CUSIP", maturity_column="Maturity Date",
                    coupon_column="Coupon Rate", rate_type_column="CPN_TYP",
                    wal_column="WAL", floor_column="CPN_Floor", header_row=1,
                ),),
            ),
        ),
    )
    flat = {q: 0.0400 for q in PROJECTION_QUARTERS}
    scenario = IncomeScenarioPaths(
        scenario_id="synthetic_demo",
        usd_3m_treasury={0: 0.0300, **{q: 0.0300 for q in PROJECTION_QUARTERS}},
        usd_1y_treasury=flat,
        usd_10y_treasury={q: 0.0450 for q in PROJECTION_QUARTERS},
        prime_rate={q: 0.0700 for q in PROJECTION_QUARTERS},
        mortgage_rate={q: 0.0650 for q in PROJECTION_QUARTERS},
    )
    return config, scenario


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", default=None, help="company-local config with [firm_data.securities]")
    parser.add_argument("--scenario", default=None, help="scenario id (company runs)")
    parser.add_argument("--warnings", choices=("summary", "full"), default="summary",
                        help="terminal warning verbosity (tables always print first)")
    parser.add_argument("--report", default=None,
                        help="write the FULL report (tables + every warning) to this file")
    parser.add_argument("--paths-out", default=None,
                        help="write the quarterly component income paths (USD millions, wide CSV: "
                             "component,PQ1..PQ9) for run_nii.py --component-paths")
    parser.add_argument("--paths-basis", choices=("xr", "full"), default="xr",
                        help="paths-out income basis: 'xr' excludes reinvestment income "
                             "(PID-SEC-8 — the reference components sheet's II_PQ columns exclude "
                             "it, so the trading residual should subtract the same basis); "
                             "'full' includes it (only if the workbook's 'Implied FRB results' "
                             "formula subtracts reinvestment too)")
    args = parser.parse_args(argv)

    if args.config is None:
        config, scenario = _synthetic_demo()
        print("SYNTHETIC DEMO (generated workbook — invented securities)")
    else:
        config = load_config(args.config)
        if config.mev is None:
            raise SystemExit("config has no [mev] section")
        scenario_id = args.scenario or next(iter(config.mev.scenarios))
        scenario = load_mev_scenario(config, scenario_id).interest_income_scenario_paths()

    inputs = load_securities_inputs(config)
    floor_mode = config.firm_data.securities.floor_mode
    floating_projection = config.firm_data.securities.floating_projection
    projection_overrides = config.firm_data.securities.floating_projection_overrides
    book_yield_categories = config.firm_data.securities.book_yield_categories
    on_error = config.firm_data.securities.on_security_error

    header = (f"firm: {inputs.firm_id}   scenario: {scenario.scenario_id}   "
              f"floor_mode: {floor_mode}   floating_projection: {floating_projection}   "
              f"overrides: {dict(projection_overrides)}   "
              f"book_yield_categories: {list(book_yield_categories)}   "
              f"maturity_source: {config.firm_data.securities.maturity_source}   "
              f"zcb_no_accretion: {list(config.firm_data.securities.zcb_no_accretion_categories)}   "
              f"a42_collapsed: {list(config.firm_data.securities.a42_collapsed_categories)}\n"
              f"maturity_rounding: {config.firm_data.securities.maturity_quarters_rounding} "
              f"overrides: {dict(config.firm_data.securities.maturity_quarters_rounding_overrides)}   "
              f"missing_ac_mode: {config.firm_data.securities.missing_ac_mode} "
              f"overrides: {dict(config.firm_data.securities.missing_ac_mode_overrides)}   "
              f"agency_face_path_survival: {config.firm_data.securities.agency_face_path_survival}   "
              f"on_security_error: {on_error}\n"
              f"AMOUNTS: USD MILLIONS per quarter — canonical unit, D-006; pre-hedge")

    results = [
        project_ust(inputs.ust, scenario, firm_id=inputs.firm_id, on_error=on_error),
        project_mbs(inputs.mbs, scenario, firm_id=inputs.firm_id, floor_mode=floor_mode, on_error=on_error,
                    reinvest_paydowns=config.firm_data.securities.reinvest_paydowns,
                    floating_projection=floating_projection, projection_overrides=projection_overrides,
                    agency_face_path_survival=config.firm_data.securities.agency_face_path_survival),
        project_other_sec(inputs.other_sec, scenario, firm_id=inputs.firm_id, floor_mode=floor_mode, on_error=on_error,
                          floating_projection=floating_projection, book_yield_categories=book_yield_categories,
                          projection_overrides=projection_overrides,
                          zcb_no_accretion_categories=config.firm_data.securities.zcb_no_accretion_categories,
                          a42_collapsed_categories=config.firm_data.securities.a42_collapsed_categories),
    ]
    tables = "\n".join(_render_model(result) for result in results)
    all_warnings = [f"loader: {w}" for w in inputs.warnings]
    for result in results:
        all_warnings.extend(f"{result.model_id}: {w}" for w in result.warnings)

    # Income tables FIRST (so they survive terminal scrollback), warnings after.
    print(header)
    print(tables)
    if args.warnings == "full":
        print(f"\nALL WARNINGS ({len(all_warnings)}):")
        for warning in all_warnings:
            print(f"  {warning}")
    else:
        print(_warning_summary(all_warnings, show=5))

    if args.report:
        with open(args.report, "w", encoding="utf-8") as handle:
            handle.write(header + "\n" + tables + f"\n\nALL WARNINGS ({len(all_warnings)}):\n")
            handle.writelines(f"  {w}\n" for w in all_warnings)
        print(f"\nfull report written to {args.report}")

    if args.paths_out:
        rows = ["component," + ",".join(f"PQ{q}" for q in PROJECTION_QUARTERS)]
        for result in results:
            path = dict(result.income_path())
            if args.paths_basis == "xr":
                for q in PROJECTION_QUARTERS:
                    path[q] -= result.quarters[q - 1].diagnostics.reinvested_income
            rows.append(result.model_id + "," + ",".join(f"{path[q]:.6f}" for q in PROJECTION_QUARTERS))
        Path(args.paths_out).write_text("\n".join(rows) + "\n", encoding="utf-8")
        print(f"\ncomponent paths written to {args.paths_out} (USD millions; basis: "
              f"{args.paths_basis}) — feed to run_nii.py --component-paths; "
              f"carries firm amounts, keep it local")


if __name__ == "__main__":
    main()
