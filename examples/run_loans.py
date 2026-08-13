"""Corporate wholesale loans (`ii_loans` — Corporate), driven by the firm workbook.

    PYTHONPATH=src python3 examples/run_loans.py                            # synthetic demo
    PYTHONPATH=src python3 examples/run_loans.py \
        --config config/local/company.toml --report loans_report.txt        # company run

Everything physical lives in the config's [firm_data.loans] section — workbook
path, every sheet name, header row and column header, the launch point, and the
scenario name (see config/company.template.toml). Nothing is passed as a CLI
path. --scenario and --launch-point exist only as one-off OVERRIDES of the
config values.

Scenario names: history rows are matched under `Actual` (through PQ0); projection
rows under the configured scenario, default `Supervisory Severely Adverse` — the
sheet's own block name. Projection rows are mapped to PQ1..PQ9 BY THEIR Date
column relative to the launch point, never by sheet order; a wrong scenario
spelling errors with the row names actually found.

Censuses print FIRST, results after — the securities-loop lesson: the diagnostics
are the product of a first run, the numbers only matter once the censuses are
clean. Output amounts: USD MILLIONS per quarter (D-006). The report file carries
firm amounts — keep it local, never commit it (gitignored `loans_report*.txt`).

The run covers both wholesale parts: Corporate always, and CRE when the config
names the H.2 sheet (`cre_h2_sheet` — PID-LOAN-18..25; the two parts use
DIFFERENT constructions on evidence, so neither borrows the other's engine).

The Federal Reserve model is PROPOSED for the 2026 stress test, NOT adopted.
Decisions: PID-LOAN-1..25 (`handbook/open-questions.md`); computation:
`specifications/interest-income/loans/ii_loans_corporate.spec.md` and
`specifications/interest-income/loans/ii_loans_cre.spec.md`."""

from __future__ import annotations

import argparse
import datetime as dt
import tempfile
from pathlib import Path

from scb_ppnr.core.schemas import PROJECTION_QUARTERS, ValidationFailure
from scb_ppnr.ingestion.config import format_effective_config, load_config
from scb_ppnr.ingestion.loans_cre_mapping import (
    CRE_CATEGORY_NAMES,
    cre_scalars_by_category_name,
)
from scb_ppnr.ingestion.loans_loader import (
    LoansSheetSpec,
    load_3m_treasury,
    load_category_balances,
    load_cre_facilities,
    load_cre_side_balances,
    load_facilities,
    load_merged_bucket_balance,
    load_reference_results,
)
from scb_ppnr.ingestion.loans_mapping import (
    DEPOSITORY_INSTITUTION_H1_CODES,
    FED_CATEGORY_NAMES,
    scalars_by_category_name,
)
from scb_ppnr.interest_income.loans_launchpoint import (
    build_cre_launch_point,
    build_launch_point,
    merged_bucket_launch_point,
)
from scb_ppnr.interest_income.loans_projection import project_corporate
from scb_ppnr.interest_income.loans_schemas import projection_quarter_index


def _run(spec: LoansSheetSpec, scenario: str, launch_point: str,
         floor_collapse: str = "balance_weighted", apply_scalar: bool = True,
         share_basis: str = "committed", balance_source: str = "m1",
         engine: str = "pid", cre_orig_date_statistic: str = "weighted_median",
         collected: list[str] | None = None) -> str:
    """Each section PRINTS the moment it is produced (the securities-loop
    lesson): a failure deep in the run must never hide the censuses that
    explain it. `collected` receives the sections for the report file even
    when a later step raises."""
    sections: list[str] = collected if collected is not None else []

    def emit(text: str) -> None:
        print(text, end="\n\n", flush=True)
        sections.append(text)

    quarters = PROJECTION_QUARTERS

    facilities, load_census = load_facilities(spec)
    balances, side_balances, m1_census = load_category_balances(spec)
    merged_balance, merged_parts = load_merged_bucket_balance(spec)
    history, base_path, launch_3m = load_3m_treasury(spec, scenario, quarters, launch_point)

    emit(
        f"RUN SETTINGS\n"
        f"  workbook      : {spec.workbook}\n"
        f"  scenario      : {scenario}\n"
        f"  launch point  : {launch_point} (PQ0); PQ1..PQ9 follow\n"
        f"  3M at PQ0     : {launch_3m:.4%}   history quarters: {len(history)} "
        f"(earliest {min(history)})\n"
        f"  engine: {engine}   floor collapse: {floor_collapse}   share basis: {share_basis}"
        f"   balance source: {balance_source}\n"
        f"  units         : USD millions; annualized decimal rates"
    )
    emit(
        "SCENARIO 3M PATH\n  PQ0 " + f"{launch_3m:7.4%}  "
        + "  ".join(f"PQ{q} {base_path[q]:7.4%}" for q in quarters)
    )
    emit(load_census.render())
    if m1_census.warnings:
        emit("M.1 NOTES\n" + "\n".join(f"  {note}" for note in m1_census.warnings))

    m1_lines = ["M.1 CATEGORY BALANCES (cross-check: the FRB SCALARS sheet's 'Sch M bal' column)"]
    for name, value in sorted(balances.items(), key=lambda kv: -kv[1]):
        m1_lines.append(f"  {value:>12,.1f}  {name}")
    m1_lines.append(
        f"  {merged_balance:>12,.1f}  merged 9/10/11 bucket from FR Y-9C "
        f"({', '.join(f'{k} {v:,.1f}' for k, v in sorted(merged_parts.items()))})"
    )
    emit("\n".join(m1_lines))

    launch, launch_diagnostics = build_launch_point(
        facilities, balances, launch_3m, history, quarters,
        lambda when: projection_quarter_index(when, launch_point),
        share_measure=share_basis,
        floor_collapse=floor_collapse,
        balance_source=balance_source,
        engine=engine,
        side_balances=side_balances,
    )
    if engine == "reference":
        # The reference bucket's balance is the M.1 sum of categories 9/10/11
        # (92,637-style), not the FR Y-9C MDRMs; its donor pool is the float
        # pool convention, which includes MIXED rows.
        m1_merged = sum(
            balances.get(FED_CATEGORY_NAMES[c], 0.0) for c in (9, 10, 11)
        )
        if m1_merged > 0.0:
            emit(
                f"MERGED BUCKET BALANCE (reference engine): {m1_merged:,.1f} from M.1 "
                f"categories 9/10/11 (FR Y-9C alternative: {merged_balance:,.1f})"
            )
            merged_balance = m1_merged
    merged = merged_bucket_launch_point(
        facilities, merged_balance, launch_3m, DEPOSITORY_INSTITUTION_H1_CODES,
        FED_CATEGORY_NAMES[9], floor_collapse=floor_collapse,
        include_mixed=(engine == "reference"),
    )
    emit(launch_diagnostics.render())

    register = ["LAUNCH-POINT REGISTER (compare each line against the workbook's own cells)"]
    register.append(
        "  segment (category/LOCOM/vt)                     share%   balance"
        "   pool rate  base@(qtr)     spread    floor   sum-wt"
    )
    for key in sorted(launch, key=str):
        lp = launch[key]
        if lp.spread is not None:
            pool = f"{lp.spread.pool_rate:8.4%}"
            base_at = lp.spread.base_quarter or "PQ0"
            base = f"{lp.spread.base_rate:7.4%}@{base_at:<7}"
            spread = f"{lp.spread.spread:8.4%}"
        else:
            pool, base, spread = "       -", "        -      ", "       -"
        floor_text = f"{lp.floor:7.4%}" if lp.floor is not None else "      -"
        wt_text = (
            f"{sum(lp.reorigination_weights.values()):7.3f}"
            if lp.reorigination_weights else "      -"
        )
        register.append(
            f"  {str(key)[:46]:<46}  {lp.share:6.2%}  {lp.balance:>8,.1f}"
            f"  {pool}  {base}  {spread}  {floor_text}  {wt_text}"
        )
    register.append(
        f"  {str(merged.segment)[:46]:<46}    n/a   {merged.balance:>8,.1f}"
        f"  {merged.spread.pool_rate:8.4%}  {merged.spread.base_rate:7.4%}@PQ0    "
        f"  {merged.spread.spread:8.4%}  "
        + (f"{merged.floor:7.4%}" if merged.floor is not None else "      -") + "        -"
    )
    emit("\n".join(register))

    emit(
        f"MERGED 9/10/11 BUCKET\n"
        f"  donor pool rate (depository floating) : {merged.spread.pool_rate:.4%}\n"
        f"  spread vs 3M(PQ0)                     : {merged.spread.spread:.4%}\n"
        f"  floor                                 : "
        + (f"{merged.floor:.4%}" if merged.floor is not None else "none on file")
    )

    if apply_scalar:
        scalars, scalar_warnings = scalars_by_category_name()
    else:
        # Reference-matching runs: the workbook's own results carry NO industry
        # scalar (verified 2026-08-12 — the grand total is the plain block sum).
        scalars, scalar_warnings = {}, ()
    projections, totals, projection_diagnostics = project_corporate(
        {**dict(launch), merged.segment: merged}, base_path, quarters, scalars,
        require_scalar=apply_scalar,
    )
    emit(projection_diagnostics.render())
    for warning in scalar_warnings:
        emit(f"WARN: {warning}")

    scaling_note = "scaled" if apply_scalar else "UNSCALED — apply_scalar = false"
    result_lines = [f"PROJECTED CORPORATE LOAN INTEREST INCOME ({scaling_note}, USD millions)"]
    result_lines.append("  category / PQ                " + "".join(f"{f'PQ{q}':>9}" for q in quarters) + "      9Q")
    grand = {q: 0.0 for q in quarters}
    for name, path in sorted(totals.items(), key=lambda kv: -sum(kv[1].values())):
        for q in quarters:
            grand[q] += path[q]
        result_lines.append(
            f"  {name[:27]:<27}  " + "".join(f"{path[q]:>9,.1f}" for q in quarters)
            + f"{sum(path.values()):>9,.1f}"
        )
    result_lines.append(
        "  TOTAL                        " + "".join(f"{grand[q]:>9,.1f}" for q in quarters)
        + f"{sum(grand.values()):>9,.1f}"
    )
    result_lines.append(f"  segments projected: {len(projections)}")
    emit("\n".join(result_lines))

    if spec.results_sheet is not None:
        reference = load_reference_results(spec)
        emit(_compare(reference, projections, base_path, quarters, scalars))

    # --- CRE part (PID-LOAN-18..25) — enabled by naming the H.2 sheet --------
    if spec.cre_h2_sheet is not None:
        cre_facilities, cre_census = load_cre_facilities(spec)
        emit(cre_census.render())

        cre_sides, cre_m1_notes = load_cre_side_balances(spec)
        m1_cre = ["CRE M.1 SIDE BALANCES (PID-LOAN-20: dom E/G per row; international = sum of I/K)"]
        for (name, side), value in sorted(cre_sides.items()):
            m1_cre.append(f"  {value:>12,.1f}  {name} / {side}")
        for note in cre_m1_notes:
            m1_cre.append(f"  {note}")
        emit("\n".join(m1_cre))

        cre_launch, cre_launch_diagnostics = build_cre_launch_point(
            cre_facilities, cre_sides, launch_3m, history, quarters,
            lambda when: projection_quarter_index(when, launch_point),
            orig_date_statistic=cre_orig_date_statistic,
        )
        emit("CRE " + cre_launch_diagnostics.render())

        cre_register = [
            f"CRE LAUNCH-POINT REGISTER (orig-date statistic: {cre_orig_date_statistic} on "
            f"outstanding, PID-LOAN-22)"
        ]
        cre_register.append(
            "  segment (category/LOCOM/vt)                     share%   balance"
            "   pool rate  base@(qtr)     spread    floor   sum-wt"
        )
        for key in sorted(cre_launch, key=str):
            lp = cre_launch[key]
            if lp.spread is not None:
                pool = f"{lp.spread.pool_rate:8.4%}"
                base_at = lp.spread.base_quarter or "PQ0"
                base = f"{lp.spread.base_rate:7.4%}@{base_at:<7}"
                spread = f"{lp.spread.spread:8.4%}"
            else:
                pool, base, spread = "       -", "        -      ", "       -"
            floor_text = f"{lp.floor:7.4%}" if lp.floor is not None else "      -"
            wt_text = (
                f"{sum(lp.reorigination_weights.values()):7.3f}"
                if lp.reorigination_weights else "      -"
            )
            cre_register.append(
                f"  {str(key)[:46]:<46}  {lp.share:6.2%}  {lp.balance:>8,.1f}"
                f"  {pool}  {base}  {spread}  {floor_text}  {wt_text}"
            )
        emit("\n".join(cre_register))

        if apply_scalar:
            cre_scalars = cre_scalars_by_category_name()
        else:
            cre_scalars = {}
        cre_projections, cre_totals, cre_projection_diagnostics = project_corporate(
            cre_launch, base_path, quarters, cre_scalars, require_scalar=apply_scalar,
        )
        emit("CRE " + cre_projection_diagnostics.render())

        cre_lines = [f"PROJECTED CRE LOAN INTEREST INCOME ({scaling_note}, USD millions)"]
        cre_lines.append("  category / PQ                " + "".join(f"{f'PQ{q}':>9}" for q in quarters) + "      9Q")
        cre_grand = {q: 0.0 for q in quarters}
        for name, path in sorted(cre_totals.items(), key=lambda kv: -sum(kv[1].values())):
            for q in quarters:
                cre_grand[q] += path[q]
            cre_lines.append(
                f"  {name[:27]:<27}  " + "".join(f"{path[q]:>9,.1f}" for q in quarters)
                + f"{sum(path.values()):>9,.1f}"
            )
        cre_lines.append(
            "  TOTAL                        " + "".join(f"{cre_grand[q]:>9,.1f}" for q in quarters)
            + f"{sum(cre_grand.values()):>9,.1f}"
        )
        cre_lines.append(
            f"  segments projected: {len(cre_projections)}   (the international category is "
            f"Fed portfolios 4-6, merged — data-forced, PID-LOAN-19)"
        )
        emit("\n".join(cre_lines))

        if spec.cre_results_sheet is not None:
            cre_reference = load_reference_results(
                spec, sheet=spec.cre_results_sheet,
                category_names=CRE_CATEGORY_NAMES, include_merged=False,
            )
            emit(_compare(
                cre_reference, cre_projections, base_path, quarters, cre_scalars,
                a8_map=cre_scalars_by_category_name(),
            ))

    return "\n\n".join(sections)


def _compare(reference, projections, base_path, quarters, scalars, a8_map=None) -> str:
    """Ours (UNSCALED, per segment) against the workbook's results blocks.

    Their block Total exceeds Fixed + Variable: the workbook sums a third stream
    whose rate signature (linear in 3M, spread above the variable stream's) says
    Mixed — so theirs' mixed is derived as Total - Fixed - Variable and compared
    to our v3 stream. Implied balance and spread per floating stream come from
    the PQ1->PQ2 slope, identically on both sides, so the decomposition into
    balance-vs-spread differences is assumption-free."""
    from scb_ppnr.interest_income.loans_schemas import (
        VT_FIXED, VT_FLOATING, VT_MIXED,
    )

    from scb_ppnr.ingestion.loans_mapping import scalars_by_category_name

    # The reference's row structure (settled 2026-08-12): Fixed and Variable rows
    # are UNSCALED, and Total = (Fixed + Variable) x the Table A8 scalar — verified
    # exact on every transcribed block, Corporate AND CRE. So: our fixed/variable
    # streams compare RAW (mixed folds into variable — the reference carries no
    # separate mixed row), and our total compares with the scalar applied.
    # `a8_map` supplies the category-name -> scalar map; default: Corporate's.
    a8 = a8_map if a8_map is not None else scalars_by_category_name()[0]
    ours: dict[tuple[str, str], dict[str, dict[int, float]]] = {}
    stream_of = {VT_FIXED: "fixed", VT_FLOATING: "variable", VT_MIXED: "variable"}
    for key, projection in projections.items():
        stream = stream_of.get(key.variable_type)
        if key.locom == "MERGED":
            stream = "variable"                     # the merged bucket is one floating block
        if stream is None:
            continue
        block = ours.setdefault((key.category, key.locom), {})
        path = block.setdefault(stream, {q: 0.0 for q in quarters})
        for q in quarters:
            path[q] += projection.income_path[q]
    lines = ["LOANS COMPARE (fixed/variable UNSCALED both sides; total = ours x Table A8 vs "
             "theirs Total; ratio = ours/theirs)"]
    lines.append("  block                                     stream       PQ1 o/t          PQ2 o/t          9Q o/t            ratio")

    def fmt(pair):
        mine, theirs = pair
        return f"{mine:>8,.1f}/{theirs:>8,.1f}"

    grand_ours = {q: 0.0 for q in quarters}
    grand_theirs = {q: 0.0 for q in quarters}
    implied_lines: list[str] = []

    for key in sorted(reference):
        category, klass = key
        streams = reference[key]
        total_theirs = streams.get("total", {})
        if all(total_theirs.get(q, 0.0) == 0.0 for q in quarters):
            continue
        block_ours = ours.get(key, {})
        scalar = a8.get(category, 1.0)
        totals_ours = {
            q: (block_ours.get("fixed", {}).get(q, 0.0)
                + block_ours.get("variable", {}).get(q, 0.0)) * scalar
            for q in quarters
        }
        for q in quarters:
            grand_ours[q] += totals_ours[q]
            grand_theirs[q] += total_theirs.get(q, 0.0)

        rows = [
            ("fixed", block_ours.get("fixed", {}), streams.get("fixed", {})),
            ("variable", block_ours.get("variable", {}), streams.get("variable", {})),
            ("total*A8", totals_ours, total_theirs),
        ]
        label = f"{category[:32]}/{klass}"
        for name, mine, theirs in rows:
            nine_mine = sum(mine.get(q, 0.0) for q in quarters)
            nine_theirs = sum(theirs.get(q, 0.0) for q in quarters)
            if nine_mine == 0.0 and nine_theirs == 0.0:
                continue
            ratio = f"{nine_mine / nine_theirs:8.4f}" if nine_theirs else "     n/a"
            lines.append(
                f"  {label:<41} {name:<9}"
                f"  {fmt((mine.get(1, 0.0), theirs.get(1, 0.0)))}"
                f"  {fmt((mine.get(2, 0.0), theirs.get(2, 0.0)))}"
                f"  {fmt((nine_mine, nine_theirs))}  {ratio}"
            )
            label = ""
            if name == "variable":
                implied = []
                for side, path in (("ours", mine), ("theirs", theirs)):
                    m1, m2 = base_path[1], base_path[2]
                    if abs(m1 - m2) < 1e-9:
                        continue
                    balance = (path.get(1, 0.0) - path.get(2, 0.0)) * 4.0 / (m1 - m2)
                    if balance <= 0.0:
                        continue
                    spread = path.get(2, 0.0) * 4.0 / balance - m2
                    implied.append(f"{side} bal {balance:>11,.0f} spread {spread:7.4%}")
                if implied:
                    implied_lines.append(
                        f"      {category[:32]}/{klass} {name:<9} " + "   ".join(implied)
                    )

    nine_ours, nine_theirs = sum(grand_ours.values()), sum(grand_theirs.values())
    lines.append(
        f"  {'GRAND (blocks present on both sides)':<41} {'total':<9}"
        f"  {fmt((grand_ours[1], grand_theirs[1]))}"
        f"  {fmt((grand_ours[2], grand_theirs[2]))}"
        f"  {fmt((nine_ours, nine_theirs))}  "
        + (f"{nine_ours / nine_theirs:8.4f}" if nine_theirs else "     n/a")
    )
    lines.append("  IMPLIED FROM THE PQ1->PQ2 SLOPE (floating streams; balance in USD millions):")
    lines.extend(implied_lines)
    lines.append(
        "  note: a matching ratio with different implied balance AND spread means the two"
    )
    lines.append(
        "  offset — fix the balance first, then re-read the spread."
    )
    return "\n".join(lines)


def _retail_enabled(spec: LoansSheetSpec) -> bool:
    return any((spec.mortgage_query_sheet, spec.card_query_sheet,
                spec.auto_pivot_sheet, spec.oc_sheet))


def _run_retail(spec: LoansSheetSpec, scenario: str, launch_point: str,
                collected: list[str] | None = None) -> str:
    """The four retail families (PID-LOAN-26..34), censuses FIRST per family.

    Each family runs iff its sheet is configured, so the mortgage slice can be
    exercised before the auto pivot file exists. Everything prints the moment
    it is produced — a failure in one family never hides another's censuses."""
    from scb_ppnr.ingestion.retail_loader import (
        load_auto_pivot,
        load_card_query,
        load_line_item_rates,
        load_mev_series,
        load_mortgage_query,
        load_oc_products,
        load_retail_m1,
    )
    from scb_ppnr.interest_income.loans_retail import (
        RetailDiagnostics,
        build_auto,
        build_card,
        build_mortgage,
        build_other_consumer,
        family_summary,
        parse_auto_scalar,
    )

    sections: list[str] = collected if collected is not None else []

    def emit(text: str) -> None:
        print(text, end="\n\n", flush=True)
        sections.append(text)

    quarters = PROJECTION_QUARTERS
    emit("RETAIL RUN — families per configured sheets; M.1 role labels are the wiring "
         "(PID-LOAN-26); scalars per PID-LOAN-32")

    m1, m1_census = load_retail_m1(spec)
    emit(m1_census.render())

    _, prime_path, prime_launch = load_mev_series(
        spec, spec.mev_prime_column, scenario, quarters, launch_point
    )
    emit("SCENARIO PRIME PATH\n  PQ0 " + f"{prime_launch:7.4%}  "
         + "  ".join(f"PQ{q} {prime_path[q]:7.4%}" for q in quarters))

    blocks = []
    diagnostics = RetailDiagnostics()

    def emit_blocks(family_blocks) -> None:
        for block in family_blocks:
            lines = [f"{block.family.upper()} BLOCK {block.name}  (scalar x{block.scalar:.4f})"]
            lines.append(f"  {'stream':<26}{'balance mm':>14}{'launch':>9}{'spread':>9}"
                         f"{'PQ1 rate':>10}{'PQ9 rate':>10}{'9Q income':>14}")
            for s in block.streams:
                launch = "" if s.launch_rate is None else f"{s.launch_rate:8.4%}"
                spread = "" if s.spread is None else f"{s.spread:8.4%}"
                pq1 = "" if s.rate_path is None else f"{s.rate_path[1]:9.4%}"
                pq9 = "" if s.rate_path is None else f"{s.rate_path[9]:9.4%}"
                lines.append(f"  {s.name:<26}{s.balance:>14,.1f}{launch:>9}{spread:>9}"
                             f"{pq1:>10}{pq9:>10}{s.total_income:>14,.2f}")
            totals = block.total_path(quarters)
            lines.append("  Total x scalar: " + "  ".join(f"PQ{q} {totals[q]:,.1f}" for q in quarters))
            emit("\n".join(lines))
        blocks.extend(family_blocks)

    if spec.mortgage_query_sheet is not None:
        query = load_mortgage_query(spec)
        emit(query.census.render())
        _, mortgage_path, mortgage_launch = load_mev_series(
            spec, spec.mev_mortgage_column, scenario, quarters, launch_point
        )
        emit("SCENARIO MORTGAGE-RATE PATH\n  PQ0 " + f"{mortgage_launch:7.4%}  "
             + "  ".join(f"PQ{q} {mortgage_path[q]:7.4%}" for q in quarters))
        emit_blocks(build_mortgage(
            query, m1,
            base_paths={"mortgage_rate": mortgage_path, "prime_rate": prime_path},
            base_launch={"mortgage_rate": mortgage_launch, "prime_rate": prime_launch},
            quarters=quarters, diagnostics=diagnostics,
        ))

    if spec.card_query_sheet is not None:
        card_segments, card_census = load_card_query(spec)
        emit(card_census.render())
        emit_blocks(build_card(
            card_segments, m1, prime_path, prime_launch,
            spec.card_spread_mode, quarters, diagnostics,
        ))

    if spec.auto_pivot_sheet is not None:
        auto = load_auto_pivot(spec)
        emit(auto.census.render())
        emit_blocks([build_auto(
            auto, m1, prime_path, prime_launch,
            parse_auto_scalar(spec.retail_auto_scalar), quarters, diagnostics,
        )])

    if spec.oc_sheet is not None:
        products, oc_census = load_oc_products(spec)
        emit(oc_census.render())
        line_rates, line_census = load_line_item_rates(spec)
        emit(line_census.render())
        emit("LINE-ITEM PQ0 RATES (jump-off earned rates, PID-LOAN-30)\n"
             + "\n".join(f"  {key:<14} {rate:8.4%}" for key, rate in sorted(line_rates.items())))
        emit_blocks(build_other_consumer(
            products, line_rates, m1, prime_path, prime_launch, quarters, diagnostics,
        ))

    emit(diagnostics.render("RETAIL DIAGNOSTICS"))

    summary = family_summary(blocks, quarters)
    lines = ["RETAIL SUMMARY (scaled; the results-summary shape: 4Q cum / 9Q cum)",
             f"  {'family':<18}{'4Q cum mm':>14}{'9Q cum mm':>14}{'4Q $bn':>9}{'9Q $bn':>9}"]
    total4 = total9 = 0.0
    for family, sums in sorted(summary.items()):
        total4 += sums["cum_4q"]
        total9 += sums["cum_9q"]
        lines.append(f"  {family:<18}{sums['cum_4q']:>14,.1f}{sums['cum_9q']:>14,.1f}"
                     f"{sums['cum_4q'] / 1e3:>9,.1f}{sums['cum_9q'] / 1e3:>9,.1f}")
    lines.append(f"  {'TOTAL RETAIL':<18}{total4:>14,.1f}{total9:>14,.1f}"
                 f"{total4 / 1e3:>9,.1f}{total9 / 1e3:>9,.1f}")
    emit("\n".join(lines))
    return "\n\n".join(sections)


def _synthetic_workbook(directory: Path) -> Path:
    """A tiny invented book so the runner is demonstrable without any .xlsx on disk."""
    import openpyxl

    book = openpyxl.Workbook()
    h1 = book.active
    h1.title = "CORP H.1"
    headers = [
        "Customer ID", "Line Reported on FR Y9C", "Interest Rate Variability",
        "Lower of Cost or Market Flag", "Interest Rate", "Committed Exposure Global",
        "Utilized Exposure Global", "Interest Rate Floor", "Origination Date", "Maturity Date",
    ]
    for _ in range(3):
        h1.append([None] * len(headers))
    h1.append(headers)
    for row in (
        ["C1", 4, 2, 3, 0.061, 400e6, 300e6, 0.02, "15-Oct-2024", "15-Oct-2030"],
        ["C2", 5, 1, 3, 0.070, 300e6, 300e6, None, "15-Jan-2020", "15-Feb-2026"],
        ["C3", 4, 3, 3, 0.055, 100e6, 80e6, None, "17-May-2022", "15-Oct-2028"],
        ["D1", 1, 2, 3, 0.055, 200e6, 150e6, None, "15-Oct-2024", "15-Oct-2027"],
        ["N1", 7, 2, 3, 0.090, 150e6, 150e6, None, "15-Oct-2024", "15-Oct-2027"],
        ["X1", 8, "[NULL]", 3, None, 50e6, 50e6, None, None, None],
        ["F1", 4, 4, 3, 0.030, 80e6, 10e6, None, "15-Oct-2024", "15-Oct-2027"],
    ):
        h1.append(row)

    fry9c = book.create_sheet("FR-Y9C 4Q 2024")
    for _ in range(7):
        fry9c.append([None, None, None])
    fry9c.append(["ID_RSSD", "Description", "Value"])
    fry9c.append(["BHCK1545", "purch/carry securities", 1_500_000.0])
    fry9c.append(["BHDM1420", "farmland", 500_000.0])

    m1 = book.create_sheet("M.1 Balance")
    for _ in range(10):
        m1.append([None] * 11)
    m1.append(["Wholesale - Corp - C&I and others", "Wholesale - Corp - C&I and others",
               "2.a Graded", "M", 700.0, "M", 90.0, "M", 180.0, "M", 10.0])
    m1.append(["Wholesale - Corp - FI", "Wholesale - Corp - FI",
               "5.d", "M", 260.0, "M", 0.0, "M", 90.0, "M", 0.0])
    # CRE rows (PID-LOAN-20 shape): E/G = domestic HFI / HFS-FVO of the row's own
    # category; I/K feed the merged international category across all three rows.
    m1.append(["Wholesale - CRE - construction", "Wholesale - CRE - international",
               "1.b(1)", "M", 210.0, "M", 20.0, "M", 15.0, "M", 0.0])       # row 13
    m1.append(["Wholesale - CRE - multi fam", "Wholesale - CRE - international",
               "1.b(2)", "M", 380.0, "M", 0.0, "M", 40.0, "M", 5.0])        # row 14
    m1.append(["Wholesale - CRE - non owner occupied", "Wholesale - CRE - international",
               "1.b(3)(b)", "M", 0.0, "M", 110.0, "M", 30.0, "M", 0.0])     # row 15

    cre = book.create_sheet("CRE H.2")
    cre_headers = [
        "Customer ID", "Line Reported on FR Y-9C", "Interest Rate Variability",
        "Lower of Cost or Market Flag", "Interest Rate", "Committed Balance",
        "Outstanding Balance", "Interest Rate Floor", "Origination Date", "Maturity Date",
    ]
    for _ in range(3):
        cre.append([None] * len(cre_headers))
    cre.append(cre_headers)
    for row in (
        # dom construction (codes 1 + 2 merge, PID-LOAN-19), floating; R1 carries
        # a 5% floor that binds once the 3M path collapses
        ["R1", 1, 2, 3, 0.070, 200e6, 150e6, 0.05, "15-Oct-2024", "15-Oct-2027"],
        ["R2", 2, 2, 3, 0.065, 100e6, 50e6, None, "15-Jan-2024", "15-Jan-2028"],
        # dom multifamily: a fixed row maturing at PQ5 and a mixed sibling on the
        # hybrid spread (PID-LOAN-23)
        ["R3", 3, 1, 3, 0.050, 300e6, 300e6, None, "15-May-2021", "15-Feb-2026"],
        ["R4", 3, 3, 3, 0.055, 80e6, 60e6, None, "17-May-2022", "15-Oct-2028"],
        # dom non-owner-occupied, HFS/FVO side
        ["R5", 5, 2, 1, 0.080, 120e6, 100e6, None, "15-Oct-2023", "15-Oct-2027"],
        # international (code 7 = Fed portfolios 4-6, merged)
        ["R6", 7, 2, 3, 0.090, 90e6, 70e6, None, "15-Oct-2023", "15-Oct-2027"],
        # DO-NOT-USE line code 4: outside every CRE category, excluded + censused
        ["R7", 4, 2, 3, 0.060, 50e6, 40e6, None, None, None],
        # fee-based and DO-NOT-USE variable types: balance-only (PID-LOAN-24)
        ["R8", 1, 4, 3, 0.030, 40e6, 30e6, None, None, None],
        ["R9", 1, "[NULL]", 3, None, 10e6, 5e6, None, None, None],
    ):
        cre.append(row)

    mev = book.create_sheet("MEV")
    mev.append(["Scenario Name", "Date", "3-month Treasury rate"])
    history = {1976: 4.9, 2020: 1.5, 2021: 0.3, 2022: 0.8}
    for year, rate in history.items():
        for quarter in (1, 2, 3, 4):
            mev.append(["Actual", f"{year} Q{quarter}", rate])
    mev.append(["Actual", "2024 Q4", 4.4])
    path_3m = [3.0, 1.8, 0.6, 0.1, 0.1, 0.1, 0.1, 0.1, 0.1]
    for index, rate in enumerate(path_3m):
        year, quarter = 2025 + index // 4, index % 4 + 1
        mev.append(["Supervisory Severely Adverse", f"{year} Q{quarter}", rate])

    target = directory / "synthetic_loans.xlsx"
    book.save(target)
    return target


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--config", type=Path, default=None,
                        help="company config with a [firm_data.loans] section; "
                             "omit for the synthetic demo")
    parser.add_argument("--scenario", default=None,
                        help="override the configured MEV projection-block name for one run")
    parser.add_argument("--launch-point", default=None,
                        help="override the configured PQ0 quarter for one run")
    parser.add_argument("--report", type=Path, default=None,
                        help="also write the output to this file (keep it local)")
    parser.add_argument("--retail-only", action="store_true",
                        help="skip the wholesale run and execute only the retail families "
                             "(their sheets must be configured in [firm_data.loans])")
    args = parser.parse_args(argv)

    collected: list[str] = []
    try:
        if args.config is None:
            with tempfile.TemporaryDirectory() as scratch:
                workbook = _synthetic_workbook(Path(scratch))
                print("SYNTHETIC DEMO — invented data, hand-checkable numbers. "
                      "Point --config at the company config for a real run.\n", flush=True)
                _run(
                    LoansSheetSpec(
                        workbook=workbook,
                        cre_h2_sheet="CRE H.2",
                        cre_m1_construction_row=13,
                        cre_m1_multifamily_row=14,
                        cre_m1_non_owner_occupied_row=15,
                    ),
                    args.scenario or "Supervisory Severely Adverse",
                    args.launch_point or "2024Q4",
                    collected=collected,
                )
        else:
            config = load_config(args.config)
            if config.firm_data is None or config.firm_data.loans is None:
                raise ValidationFailure(
                    f"{args.config}: no [firm_data.loans] section — the loans run is configured "
                    f"there (workbook path, sheet names, launch point; see "
                    f"config/company.template.toml)"
                )
            loans = config.firm_data.loans
            print(format_effective_config(config) + "\n", flush=True)
            if args.retail_only and not _retail_enabled(loans.spec):
                raise ValidationFailure(
                    "--retail-only was passed but no retail sheet is configured "
                    "(mortgage_query_sheet / card_query_sheet / auto_pivot_sheet / oc_sheet)"
                )
            if not args.retail_only:
                _run(
                    loans.spec,
                    args.scenario or loans.scenario,
                    args.launch_point or loans.launch_point,
                    loans.floor_collapse,
                    loans.apply_scalar,
                    loans.share_basis,
                    loans.balance_source,
                    loans.engine,
                    loans.cre_orig_date_statistic,
                    collected=collected,
                )
            if _retail_enabled(loans.spec):
                _run_retail(
                    loans.spec,
                    args.scenario or loans.scenario,
                    args.launch_point or loans.launch_point,
                    collected=collected,
                )
        failed = None
    except ValidationFailure as error:
        failed = str(error)
        print("\nVALIDATION FAILURE — the run stopped here; every section above "
              "already printed:\n  " + failed, flush=True)

    if args.report is not None:
        body = "\n\n".join(collected)
        if failed is not None:
            body += f"\n\nVALIDATION FAILURE\n  {failed}"
        args.report.write_text(body + f"\n\ngenerated {dt.datetime.now():%Y-%m-%d %H:%M}\n",
                               encoding="utf-8")
        print(f"\nreport written to {args.report} — carries firm amounts; keep it local")
    return 0 if failed is None else 1


if __name__ == "__main__":
    raise SystemExit(main())
